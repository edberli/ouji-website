#!/usr/bin/env python3
"""
Build and publish the fwee range.

fwee's rows are the messiest of any brand: the shade code sits inline
with no consistent separator ("… Pudding Pot BS", "… Pudding Pot- MV05",
"*Keyring* … CR", and one title with a double space), so a title-prefix
split produced 57 groups out of 98 rows. Lines are identified by price
plus a series keyword instead, and the shade is whatever survives after
the series name is stripped.

Shades, barcodes, prices and stock are read from the workbook at build
time — never transcribed.

    python3 scripts/build_fwee.py mirror
    python3 scripts/build_fwee.py publish [--dry-run]
"""
import re

import openpyxl

from brand_build import run

VENDOR = "fwee"
STOCK_XLSX = "/Volumes/core/下載/cosmetic_products_shopify.xlsx"

T_LIP = "fwee, K-Beauty, 彩妝, 唇妝, makeup, lip"
T_CHEEK = "fwee, K-Beauty, 彩妝, 修容, makeup, cheek"
T_EYE = "fwee, K-Beauty, 彩妝, 眼妝, makeup, eye"
T_BASE = "fwee, K-Beauty, 彩妝, 底妝, makeup, base"
SPECS = ["產地：韓國 Made in Korea"]

# slug -> (title, type, tags, price, must-match /must-not-match keyword, copy)
LINES = [
    ("fwee-lip-cheek-blurry-pudding-pot", "fwee Lip&Cheek Blurry Pudding Pot 布丁唇頰泥",
     "胭脂", T_CHEEK + ", blush", 128, (None, "HIGHLIGHT"), dict(
        hook="fwee 賣到斷貨嗰盒布丁。",
        lede="布丁質地一撳落去會彈返上嚟，指腹輕拍就融入底妝，唔會浮喺粉底上面。唇同頰用同一盒，色調自動夾——出門帶一盒就夠。",
        bullets=[("布丁彈潤質地", "貼膚唔浮，唔會結塊。"),
                 ("唇頰兩用", "一盒搞掂全臉氣色。"),
                 ("柔霧收尾", "唔油亮，唔會過分啞。"),
                 ("色域最闊", "全線最多色，冷暖膚色都揀到。")],
        how="指腹沾取少量，點於顴骨後由內向外輕拍；點唇時由中央推開。")),

    ("fwee-lip-cheek-glowy-jelly-pot", "fwee Lip&Cheek Glowy Jelly Pot 果凍唇頰凍",
     "胭脂", T_CHEEK + ", blush", 118, (None, None), dict(
        hook="布丁嘅水光版本。",
        lede="同 Pudding Pot 同一個概念，但收尾係水光而唔係柔霧。果凍質地薄透，上臉似皮膚本身透出嚟嘅光——唔會有胭脂嘅粉感。",
        bullets=[("果凍水光", "透薄有光澤，唔黐。"),
                 ("唇頰兩用", "色調自動統一。"),
                 ("融入底妝", "唔會推花粉底。"),
                 ("可疊可薄", "薄一層氣色，多兩層打卡妝。")],
        how="指腹沾取米粒大小，點於顴骨後輕拍暈開。")),

    ("fwee-3d-volumizing-glass-tint", "fwee 3D 立體玻璃唇釉", "唇釉", T_LIP + ", liptint",
     108, (None, None), dict(
        hook="唇形立體咗，唔使打唇部高光。",
        lede="玻璃級光澤集中喺唇中央，自然做出立體感——即係話唔使另外用高光去堆唇珠。薄透唔黐，戴口罩都放心。",
        bullets=[("3D 立體光澤", "唇形自然更飽滿。"),
                 ("玻璃質感", "反光度高但唔黏。"),
                 ("薄透可疊", "單搽裸唇，疊喺唇膏上變水光。"),
                 ("色域齊全", "由裸調到莓紅。")],
        how="沿唇形塗抹，集中點唇中央做立體效果。")),

    ("fwee-glitz-stone-highlighter", "fwee Glitz Stone Highlighter 寶石高光", "高光",
     T_CHEEK + ", highlighter", 128, ("HIGHLIGHT", None), dict(
        hook="十二隻色，由自然氣色到派對閃。",
        lede="一般高光只有香檳同珍珠白。呢個系列十二隻色由柔和裸光一路去到彩色寶石光，日常同派對都搵到啱嘅——而且珠光細，唔會一粒粒。",
        bullets=[("十二色最闊", "由自然到大膽，全線最齊。"),
                 ("細緻珠光", "唔見閃片顆粒。"),
                 ("唔卡乾紋", "細滑貼膚。"),
                 ("可疊加", "薄一層氣色，多兩層上鏡。")],
        how="以刷具點於顴骨最高點、眉骨同鼻樑，再向外推開。")),

    ("fwee-pocket-eye-palette", "fwee Pocket 迷你便攜五色眼影盤", "眼影",
     T_EYE + ", eyeshadow, palette", 148, (None, None), dict(
        hook="銀包大細，五格夠用。",
        lede="大盤帶唔出街，細盤又通常唔夠色。呢個五格已經齊晒打底、暈染、加深同閃片，順住用就係完整眼妝，而且袋得落銀包。",
        bullets=[("五格夠用", "打底、暈染、加深、閃片。"),
                 ("袋裝尺寸", "化妝袋唔佔位。"),
                 ("配色已諗好", "順住次序落就得。"),
                 ("粉質細滑", "唔飛粉，唔會落眼底。")],
        how="淺色打底整個眼窩，中間色暈染褶位，深色壓眼尾，閃片點眼中央。")),

    ("fwee-pocket-cheek-palette", "fwee Pocket 迷你便攜修容盤", "修容",
     T_CHEEK + ", palette", 168, (None, None), dict(
        hook="修容、腮紅、高光，一盒帶得走。",
        lede="出門補妝最麻煩係要開三個盒。呢盤把修容、腮紅同高光放埋一齊，色調事先夾好，唔使自己諗點配。",
        bullets=[("三格一盒", "修容、腮紅、高光一次過。"),
                 ("色調已配好", "同盤互相夾，唔會撞色。"),
                 ("袋裝尺寸", "旅行同補妝最啱。"),
                 ("粉質細滑", "唔會浮粉，唔會結塊。")],
        how="深色收修輪廓，中間色掃顴骨，最後點高光於顴骨最高點。")),

    ("fwee-spa-glowing-uv-tone-up-base", "fwee SPA 光澤 UV 提亮妝前乳霜", "底妝",
     T_BASE + ", primer", 138, (None, None), dict(
        hook="提亮、打底、防曬，一步做齊。",
        lede="早上趕時間最想慳步驟。呢款妝前乳同時做提亮、平滑底妝同 UV 防護，搽完直接上粉底就得——而且提亮係自然透亮唔係死白。",
        bullets=[("三合一", "提亮、打底、防曬。"),
                 ("唔死白", "自然透亮。"),
                 ("平滑底妝", "粉底更均勻更貼。"),
                 ("兩種收尾", "柔光同水光可選。")],
        how="潔面保養後取適量薄塗全臉，待吸收後再上底妝。")),

    ("fwee-one-minute-ready-lip-serum", "fwee 瞬間水潤唇部精華", "唇彩",
     T_LIP + ", lipbalm", 88, (None, None), dict(
        hook="化妝前一分鐘，唇就唔會起皮。",
        lede="唇乾嗰陣點搽唇膏都會卡紋。呢支唇部精華吸收快，搽完一分鐘就可以上唇膏，唔會令唇膏浮起——亦可以當夜間唇膜用。",
        bullets=[("吸收快", "一分鐘就可以上唇妝。"),
                 ("即時撫平唇紋", "唇膏唔會卡紋。"),
                 ("可當唇膜", "夜間厚敷保養。"),
                 ("透明百搭", "任何唇色都用得。")],
        how="上唇妝前薄搽一層；或睡前厚敷當唇膜。")),
]

# 18 Pudding Pot rows reached us with the shade stripped from the title.
# These came back from barcode lookups against the manufacturer's
# catalogue and from retailer listings; four others appear in no source
# we can reach and stay off the product until someone reads the box.
BY_BARCODE = {
    "8809652582498": "ND02 Like",
    "8809652582573": "CR05 Girls",
    "8809652582627": "PK05 Sth",
    "8809652582689": "MV01 Chill'N",
    "8809652582559": "CR03 BFF",
    "8809652582603": "PK03 Cherry",
    "8809652582610": "PK04 Crush",
    "8809652582658": "RD03 Ambitious",
    "8809652582665": "Fav",
    "8809652582672": "RD05 Greedy",
    "8809652582719": "MV04 Slayyy",
    "8809652582764": "RS04 Memories",
    "8809652582771": "RS05 Film",
}
# Filed under the Pudding Pot price but actually the lip serum.
REASSIGN = {"8809652585796": "fwee-one-minute-ready-lip-serum"}
SERUM_SHADE = {"8809652585796": "Keyring 套裝"}

SERIES = re.compile(
    r"^\s*FWEE\s*[-–]?\s*(\*Keyring\*)?\s*"
    r"(Lip\s*&?\s*Cheek\s+Blurry\s+Pudding\s+Pot|Lip\s*&?\s*Cheek\s+Glowy\s+Jelly\s+Pot"
    r"|GLITZ\s+STONE\s+HIGHLIGHTER|POCKET\s+EYE\s+PALETTE|POCKET\s+CHEEK\s+PALETTE"
    r"|3D立體玻璃唇釉\s*30%?|SPA光澤UV提亮妝前乳霜|瞬間水潤唇部精華|迷你便攜五色眼影盤|眼影盤)?"
    r"\s*[-–]?\s*", re.I)


def rows():
    wb = openpyxl.load_workbook(STOCK_XLSX)
    ws = wb.active
    h = [c.value for c in ws[1]]
    iT, iV, iQ, iP, iB = (h.index(x) for x in
                          ("Title", "Vendor", "Variant Inventory Qty",
                           "Variant Price", "Variant Barcode"))
    return [(str(r[iT]).strip(), str(r[iB]).strip(), r[iP], r[iQ] or 0)
            for r in ws.iter_rows(min_row=2, values_only=True)
            if r[iV] and "fwee" in str(r[iV]).lower()]


def shade_of(title):
    """None when the row carries no shade at all — 17 Pudding Pot rows are
    just the series name, and inventing labels for them would put the
    wrong thing on the box."""
    tail = SERIES.sub("", title, count=1).strip()
    tail = re.sub(r"\s*\d+(\.\d+)?\s*(g|ml)\s*$", "", tail, flags=re.I)
    tail = re.sub(r"\s{2,}", " ", tail).strip()
    return tail or None


P = {}
NO_SHADE = []
all_rows = rows()
used = set()
for slug, title, ptype, tags, price, (need, avoid), copy in LINES:
    matched = [(BY_BARCODE.get(b) or SERUM_SHADE.get(b) or shade_of(t), b, q) for t, b, p, q in all_rows
               if (REASSIGN.get(b) == slug or p == price) and b not in used
               and REASSIGN.get(b, slug) == slug
               and (REASSIGN.get(b) or not need or need.lower() in t.lower())
               and (REASSIGN.get(b) or not avoid or avoid.lower() not in t.lower())]
    used.update(b for _, b, _ in matched)
    picked = [x for x in matched if x[0]]
    NO_SHADE.extend(b for n, b, _ in matched if not n)
    if not picked:
        continue
    P[slug] = dict(title=title, type=ptype, tags=tags, price=price, specs=SPECS,
                   hook=copy["hook"], lede=copy["lede"],
                   bullets=copy["bullets"], how=copy["how"],
                   shades=sorted(set(picked)))

_missed = [t for t, b, _, _ in all_rows if b not in used]
if _missed:
    print(f"未分組（{len(_missed)}）: " + "; ".join(_missed[:5]))
if NO_SHADE:
    print(f"來源檔冇色號、暫時唔上架（{len(NO_SHADE)}）：" + ", ".join(NO_SHADE))

run(__name__, VENDOR, P, "fwee")
