#!/usr/bin/env python3
"""
Build and publish the hince range.

hince is the most awkward of the brands to group: its supplier titles run
the shade code inline with no brackets ("hince RAW GLOW GEL TINT R012
RAW FIG"), so lines are identified by their series prefix and the shade
is whatever follows it.

The brand sells a restrained, editorial look — muted mood colours and
skin that reads as skin — so the copy leans on how quiet the results are
rather than on pigment or coverage.

    python3 scripts/build_hince.py mirror
    python3 scripts/build_hince.py publish [--dry-run]
"""
import re

import openpyxl

from brand_build import run

VENDOR = "hince"
STOCK_XLSX = "/Volumes/core/下載/cosmetic_products_shopify.xlsx"

T_EYE = "hince, K-Beauty, 彩妝, 眼妝, makeup, eye"
T_LIP = "hince, K-Beauty, 彩妝, 唇妝, makeup, lip"
T_CHEEK = "hince, K-Beauty, 彩妝, 修容, makeup, cheek"
T_BASE = "hince, K-Beauty, 彩妝, 底妝, makeup, base"
SPECS = ["產地：韓國 Made in Korea"]

# slug -> (title, type, tags, series prefix in the supplier title, copy)
LINES = [
    ("hince-raw-glow-gel-tint", "hince Raw Glow Gel Tint 果凍唇釉", "唇釉", T_LIP + ", liptint",
     r"RAW\s*(GLOW\s*)?GEL TINT", dict(
        hook="似唇本身嘅顏色，只係更好睇。",
        lede="hince 嘅色調全部收斂克制，唔會有「一搽就好濃妝」嘅問題。Raw Glow 果凍質地薄透貼唇，光澤自然唔黐——十七隻色，總搵到你日常嗰隻。",
        bullets=[("果凍薄透", "貼唇唔黐，唔會黐頭髮。"),
                 ("自然光澤", "水潤但唔油亮。"),
                 ("低飽和色調", "返工、見客都用得。"),
                 ("十七色最闊", "全線色域最齊。")],
        how="沿唇形塗一層；想飽和啲就等半乾後再疊。")),

    ("hince-raw-glow-dewy-ball", "hince Raw Glow Dewy Ball 滾珠唇頰露", "胭脂",
     T_CHEEK + ", blush", r"RAW GLOW DEWY BALL", dict(
        hook="滾珠上臉，唔使工具。",
        lede="液狀唇頰露最怕出量控制唔到。滾珠每次只帶出薄薄一層，喺笑肌滾兩下再用指腹拍散，就係由皮膚透出嚟嘅血色。",
        bullets=[("滾珠控量", "唔會一次出太多。"),
                 ("唇頰兩用", "一支搞掂全臉氣色。"),
                 ("融入底妝", "唔會推花粉底。"),
                 ("十三色可選", "由玫瑰、蜜桃到莓調。")],
        how="喺笑肌滾兩下，再以指腹由內向外輕拍暈開。")),

    ("hince-mood-enhancer-water-liquid-glow", "hince Mood Enhancer Water Liquid Glow 水光唇釉",
     "唇釉", T_LIP + ", liptint", r"MOOD ENHANCER WATER LIQUID(?:\s*GLOW)?", dict(
        hook="水光，但唔會亮到搶戲。",
        lede="Mood Enhancer 系列係 hince 嘅招牌。水感質地薄如一層膜，光澤克制——想要好氣色但唔想被人一眼睇出化咗妝嗰陣最啱。",
        bullets=[("薄如水膜", "唔會有厚重唇釉嘅黏膩感。"),
                 ("克制光澤", "水潤但唔浮誇。"),
                 ("持色自然", "淡出均勻，唔會斑駁。"),
                 ("十色可選", "全部係日常用得着嘅調子。")],
        how="沿唇形塗抹，抿唇令顏色均勻。")),

    ("hince-mood-enhancer-matte", "hince Mood Enhancer Matte 霧面唇膏", "唇膏",
     T_LIP + ", lipstick", r"MOOD ENHANCER MATTE", dict(
        hook="霧面唇膏，但唔會扯到唇乾。",
        lede="一般霧面唇膏靠揮發做啞光，代價就係乾。Mood Enhancer Matte 質地綿滑，上唇柔霧但保留潤度——正式場合用得，全日都唔使補。",
        bullets=[("柔霧唔乾", "唔會起皮，唔會顯唇紋。"),
                 ("一塗即勻", "唔使唇刷，唔會結塊。"),
                 ("高持久度", "食完嘢都唔會全脫。"),
                 ("八色可選", "由裸調到深沉玫瑰。")],
        how="由唇中央向外塗抹，再以指腹拍散邊緣。")),

    ("hince-mood-enhancer-lip-glow", "hince Mood Enhancer Lip Glow 潤色唇膏", "唇膏",
     T_LIP + ", lipbalm", r"MOOD ENHANCER LIP GLOW", dict(
        hook="潤唇為主，帶色為輔。",
        lede="唔想化妝但又想唇色好啲嗰啲日子——Lip Glow 滋潤度似潤唇膏，色淡到似天生唇色，出街前隨手搽兩下就得。",
        bullets=[("極滋潤", "乾唇都唔會起皮。"),
                 ("微微帶色", "唔似化咗妝。"),
                 ("隨時補搽", "唔使照鏡都塗得均勻。"),
                 ("五色可選", "由玫瑰到裸調。")],
        how="直接塗於唇部，隨時補搽。")),

    ("hince-dewy-liquid-cheek", "hince Dewy Liquid Cheek 水感胭脂", "胭脂", T_CHEEK + ", blush",
     r"DEWY LIQUID CHEEK", dict(
        hook="唔似搽咗胭脂，似本身好氣色。",
        lede="液體質地一觸即化，指腹輕拍就融入底妝，唔會浮喺粉底上面。hince 嘅色調偏柔，所以就算落多咗手都唔會突兀。",
        bullets=[("融入底妝", "唔會推花粉底，唔會結塊。"),
                 ("免工具", "手指拍兩下就完成。"),
                 ("柔調配色", "唔會一撲就變紅蘋果。"),
                 ("五色可選", "由暖雀斑、日光蜜桃到冷調紫。")],
        how="指腹沾取米粒大小，點於顴骨後由內向外輕拍。")),

    ("hince-true-dimension-radiance-balm", "hince True Dimension Radiance Balm 水光棒", "高光",
     T_CHEEK + ", highlighter", r"(?:TRUE DIMENSION RADIANCE BALM|通透保濕透亮高光\s*水光棒\S*\s*-)", dict(
        hook="唔使刷、唔使鏡，三秒還你剛睡飽嘅臉。",
        lede="膏體遇溫即化，直接推上面就融入底妝，唔會推花、唔會斷層。光澤係「透出嚟」而唔係「貼上去」——遠睇係氣色，近睇係好皮膚。",
        bullets=[("免工具", "直接推上面，通勤途中都補得到。"),
                 ("乾肌救星", "掃過脫皮位都唔會起皮。"),
                 ("自然光感", "唔見閃片，只見光澤。"),
                 ("七色可選", "由通透、晨光到柔粉調。")],
        how="以棒頭喺顴骨、眉骨、鼻樑輕輕畫兩下，再用指腹輕印暈開。")),

    ("hince-all-round-eye-palette", "hince All-Round Eye Palette 眼影盤", "眼影",
     T_EYE + ", eyeshadow, palette", r"ALL-ROUND EYE PALETTE", dict(
        hook="一盤一種情緒，唔使諗配色。",
        lede="每盤圍繞一個主題配好色，啞光同珠光齊全，由淺到深排好——照住次序落就已經係完整妝容，新手都唔會撞色。",
        bullets=[("配色已諗好", "順住格數用就得。"),
                 ("啞光＋珠光", "質地齊全，唔使配第二盤。"),
                 ("粉質綿密", "唔飛粉，唔會落喺眼底。"),
                 ("五款調子", "由九月、深秋到玫瑰同霧藍。")],
        how="淺色打底整個眼窩，中間色暈染褶位，深色壓眼尾，珠光點眼中央。")),

    ("hince-second-skin-glow-cushion", "hince Second Skin Glow Cushion 水光氣墊粉底",
     "氣墊粉底", T_BASE + ", cushion", r"SECOND SKIN GLOW CUSHION", dict(
        hook="第二層皮膚，唔係一層粉。",
        lede="Second Skin 系列嘅名唔係講笑——薄薄一層就夠勻，妝感係健康皮膚嘅光澤而唔係粉底嘅厚度。乾肌同混合肌都撐得住。",
        bullets=[("薄塗夠遮", "唔使疊厚就均勻。"),
                 ("水光收尾", "亮而唔油，唔會似出汗。"),
                 ("唔卡粉", "乾燥位都唔會起皮。"),
                 ("四個色階", "15 到 23，白皮到自然色。")],
        how="以粉撲輕拍上臉，由面中央向外推開。")),

    ("hince-second-skin-mesh-matte-cushion", "hince Second Skin Mesh Matte Cushion 霧面氣墊粉底",
     "氣墊粉底", T_BASE + ", cushion", r"SECOND SKIN MESH MATTE CUSHION", dict(
        hook="霧面，但唔會乾到見紋。",
        lede="出油肌用水光氣墊，中午就會反光到似出汗。Mesh Matte 做柔霧收尾，控油之餘保留潤度，妝面平滑但唔會死板。",
        bullets=[("柔霧收尾", "控油唔反光。"),
                 ("平滑毛孔", "妝面均勻。"),
                 ("持妝唔氧化", "全日唔變深變黃。"),
                 ("四個色階", "15 到 23。")],
        how="以粉撲輕拍上臉；出油位可多拍一層。")),

    ("hince-second-skin-foundation", "hince Second Skin Foundation 粉底液", "底妝",
     T_BASE + ", foundation", r"SECOND SKIN FOUNDATION", dict(
        hook="遮得到，但望落仲係皮膚。",
        lede="遮瑕力同自然感通常二選一。Second Skin 粉底液質地輕薄但色素密，薄塗已經勻，唔使為咗遮瑕而搽到成塊面似面具。",
        bullets=[("薄塗高遮", "唔使疊厚。"),
                 ("自然膚感", "唔會有面具感。"),
                 ("持妝一整日", "唔氧化唔脫妝。"),
                 ("四個色階", "17 到 23。")],
        how="以粉底刷或海綿由面中央向外推開；瑕疵位再輕拍一層。")),

    ("hince-second-skin-cover-concealer", "hince Second Skin Cover Concealer 遮瑕膏", "底妝",
     T_BASE + ", concealer", r"SECOND SKIN COVER CONCEALER", dict(
        hook="遮到黑眼圈，但唔會卡紋。",
        lede="眼下遮瑕最怕落粉之後全部卡喺細紋度。呢支質地柔滑貼膚，遮蓋力夠但唔會積聚——笑起上嚟都唔會裂。",
        bullets=[("唔卡細紋", "眼下用一整日都平滑。"),
                 ("高遮蓋", "黑眼圈同痘印都遮得到。"),
                 ("易推開", "唔會推花底妝。"),
                 ("四個色階", "17 到 23。")],
        how="點於瑕疵位，以指腹或遮瑕刷由中央向外輕拍。")),

    ("hince-second-skin-tone-up-base", "hince Second Skin Tone Up Base 校色妝前乳", "底妝",
     T_BASE + ", primer", r"SECOND SKIN TONE UP BASE", dict(
        hook="校色，唔係漂白。",
        lede="一般提亮妝前乳會令膚色變得死白。呢款用綠、紫、蜜桃三種調子分別壓泛紅、暗黃同灰暗——校完再上底妝，用量少好多。",
        bullets=[("三款校色", "綠壓泛紅、紫壓暗黃、蜜桃提氣色。"),
                 ("唔死白", "自然提亮，唔會似戴面具。"),
                 ("同時打底", "平滑毛孔，底妝更貼。"),
                 ("薄塗即可", "少量已經有效。")],
        how="取少量點於需要校色嘅位置，推勻後再上底妝。")),

    ("hince-second-skin-airy-powder", "hince Second Skin Airy Powder 輕盈定妝蜜粉", "底妝",
     T_BASE + ", powder", r"SECOND SKIN AIRY POWDER", dict(
        hook="定妝，但唔會令皮膚變乾。",
        lede="蜜粉一撲就乾、就見紋，多數係粉體太粗。Airy Powder 粉體極細，薄薄一層定住底妝，摸落仲係滑嘅。",
        bullets=[("極細粉體", "唔會見紋，唔會浮粉。"),
                 ("控油唔乾", "定妝之餘保留潤度。"),
                 ("兩款色調", "薰衣草提亮、米色定妝。"),
                 ("可全臉可局部", "T 字位單獨定妝都得。")],
        how="以粉撲或蜜粉刷沾取少量，輕按於全臉或出油位。")),

    ("hince-second-skin-hydrating-primer", "hince Second Skin Hydrating Primer 保濕妝前乳",
     "底妝", T_BASE + ", primer", r"SECOND SKIN HYDRATING PRIMER", dict(
        hook="乾肌上妝前嗰一步。",
        lede="底妝卡粉、起皮，多數唔係粉底問題而係皮膚太乾。呢支妝前乳補足水分再成一層薄膜，之後上粉底就會貼服好多。",
        bullets=[("即時保濕", "乾燥位唔會再起皮。"),
                 ("平滑底妝", "粉底更均勻更貼。"),
                 ("延長持妝", "唔會中途斑駁。"),
                 ("清爽唔黏", "搽完即刻可以上妝。")],
        how="潔面保養後取適量薄塗全臉，待吸收後再上底妝。")),
]


def rows():
    wb = openpyxl.load_workbook(STOCK_XLSX)
    ws = wb.active
    h = [c.value for c in ws[1]]
    iT, iV, iQ, iP, iB = (h.index(x) for x in
                          ("Title", "Vendor", "Variant Inventory Qty",
                           "Variant Price", "Variant Barcode"))
    return [(str(r[iT]).strip(), str(r[iB]).strip(), r[iP], r[iQ] or 0)
            for r in ws.iter_rows(min_row=2, values_only=True)
            if r[iV] and "hince" in str(r[iV]).lower()]


def shade_after(title, prefix):
    """Everything after the series name, minus the trailing size."""
    m = re.search(prefix + r"\s*(.+)$", title, re.I)
    tail = m.group(m.lastindex) if m and m.lastindex else title
    return re.sub(r"\s*\d+(\.\d+)?\s*(g|ml)\s*$", "", tail, flags=re.I).strip() or title


P = {}
all_rows = rows()
used = set()
for slug, title, ptype, tags, prefix, copy in LINES:
    picked = [(shade_after(t, prefix), b, q, p) for t, b, p, q in all_rows
              if re.search(prefix, t, re.I) and b not in used]
    used.update(b for _, b, _, _ in picked)
    if not picked:
        continue
    P[slug] = dict(title=title, type=ptype, tags=tags,
                   price=max(p for _, _, _, p in picked) or 0, specs=SPECS,
                   hook=copy["hook"], lede=copy["lede"],
                   bullets=copy["bullets"], how=copy["how"],
                   shades=sorted((n, b, q) for n, b, q, _ in picked))

_missed = [t for t, b, _, _ in all_rows if b not in used]
if _missed:
    print(f"未分組（{len(_missed)}）: " + "; ".join(_missed[:6]))

run(__name__, VENDOR, P, "hince")
