#!/usr/bin/env python3
"""
Put the supplier's cost prices into Shopify, and derive the ranking the
catalogue's "推薦" order needs.

The workbook carried `Cost per item` for all 960 rows from the start —
the publisher just never wrote it, because `build_input` only set
barcode, price and quantity per variant. So the store has had no margin
data and "推薦" has been ranking on awards and price instead.

Two outputs:

  * Shopify `InventoryItem.cost`, matched by barcode. That is what makes
    Shopify's own profit reports work.
  * `featured.json` for the site — and deliberately NOT the costs. Unit
    profit is commercially sensitive; anyone could read it out of a
    static file and price against us. What ships is a 0–100 rank, which
    orders the grid without disclosing what anything earns.

    python3 scripts/push_costs.py --dry-run
    python3 scripts/push_costs.py
"""
import argparse
import json
import os
import sys

import openpyxl

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from shopify_admin import gql, user_errors  # noqa: E402

STOCK_XLSX = "/Volumes/core/下載/cosmetic_products_shopify.xlsx"
OUT = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))),
                   "featured.json")

VARIANTS = """
query($cursor: String) {
  productVariants(first: 200, after: $cursor) {
    pageInfo { hasNextPage endCursor }
    edges { node {
      id barcode price
      product { handle status }
      inventoryItem { id unitCost { amount } }
    } }
  }
}
"""

SET_COST = """
mutation($id: ID!, $input: InventoryItemInput!) {
  inventoryItemUpdate(id: $id, input: $input) {
    inventoryItem { id unitCost { amount } }
    userErrors { field message }
  }
}
"""


def costs_from_workbook():
    wb = openpyxl.load_workbook(STOCK_XLSX)
    ws = wb.active
    h = [c.value for c in ws[1]]
    iB, iC = h.index("Variant Barcode"), h.index("Cost per item")
    out = {}
    for r in ws.iter_rows(min_row=2, values_only=True):
        bar, cost = r[iB], r[iC]
        if bar and cost not in (None, ""):
            try:
                out[str(bar).strip()] = round(float(cost), 2)
            except (TypeError, ValueError):
                pass
    return out


def variants():
    cursor, out = None, []
    while True:
        d = gql(VARIANTS, {"cursor": cursor})["productVariants"]
        out += [e["node"] for e in d["edges"]]
        if not d["pageInfo"]["hasNextPage"]:
            return out
        cursor = d["pageInfo"]["endCursor"]


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--dry-run", action="store_true")
    args = ap.parse_args()

    costs = costs_from_workbook()
    print(f"workbook 有成本嘅 SKU：{len(costs)}")

    # handle -> best unit profit across its variants
    profit, wrote, missing, already = {}, 0, [], 0
    for v in variants():
        if v["product"]["status"] != "ACTIVE":
            continue
        bar = (v["barcode"] or "").strip()
        cost = costs.get(bar)
        if cost is None:
            missing.append(bar or v["product"]["handle"])
            continue
        gain = float(v["price"]) - cost
        h = v["product"]["handle"]
        profit[h] = max(profit.get(h, 0.0), gain)

        have = (v["inventoryItem"].get("unitCost") or {}).get("amount")
        if have is not None and abs(float(have) - cost) < 0.005:
            already += 1
            continue
        if args.dry_run:
            wrote += 1
            continue
        out = gql(SET_COST, {"id": v["inventoryItem"]["id"],
                             "input": {"cost": f"{cost:.2f}"}})
        user_errors(out, "inventoryItemUpdate")
        wrote += 1
        if wrote % 100 == 0:
            print(f"    寫咗 {wrote}")

    print(f"寫入成本 {wrote}；已經啱 {already}；workbook 冇對應 {len(missing)}")
    for b in missing[:8]:
        print("   ", b)

    # Rank, not money: 0–100, highest unit profit first.
    if profit:
        top = max(profit.values()) or 1
        ranks = {h: round(max(0.0, g) / top * 100) for h, g in profit.items()}
        if not args.dry_run:
            with open(OUT, "w") as f:
                json.dump({"profitRank": ranks}, f, separators=(",", ":"))
            print(f"\n寫咗 {OUT}（{len(ranks)} 件產品，只有排名冇金額）")
        best = sorted(profit.items(), key=lambda x: -x[1])[:5]
        print("\n單件毛利最高：")
        for h, g in best:
            print(f"  {g:7.2f}  {h}")


if __name__ == "__main__":
    main()
