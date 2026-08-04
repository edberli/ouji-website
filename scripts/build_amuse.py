#!/usr/bin/env python3
"""
Build and publish the AMUSE range.

AMUSE is half its own vegan line and half Sanrio collaborations (My
Melody, Kuromi), and the two need different copy: the regular line sells
on formula, the collabs sell on the packaging being a limited object.
The two groups are kept as separate products rather than merged, since a
shopper hunting the collab is not shopping for a cushion in general.

Shades are read straight from the supplier workbook rather than typed
here — transcribing them by hand once produced a full set of invented
barcodes on another brand.

    python3 scripts/build_amuse.py mirror
    python3 scripts/build_amuse.py publish [--dry-run]
"""
import collections
import re

import openpyxl

from brand_build import run

VENDOR = "AMUSE"
STOCK_XLSX = "/Volumes/core/下載/cosmetic_products_shopify.xlsx"

T_EYE = "AMUSE, K-Beauty, 彩妝, 眼妝, makeup, eye, vegan"
T_LIP = "AMUSE, K-Beauty, 彩妝, 唇妝, makeup, lip, vegan"
T_CHEEK = "AMUSE, K-Beauty, 彩妝, 修容, makeup, cheek, vegan"
T_BASE = "AMUSE, K-Beauty, 彩妝, 底妝, makeup, base, vegan"
T_BODY = "AMUSE, K-Beauty, 身體護理, body, vegan"
SPECS = ["產地：韓國 Made in Korea", "純素配方 Vegan"]

# slug -> (title, type, tags, matcher against the supplier title, copy)
LINES = [
    ("amuse-dew-tint", "AMUSE Dew Tint 露水唇釉", "唇釉", T_LIP + ", liptint",
     lambda t: "露水唇釉" in t, dict(
        hook="似含住一啖水。",
        lede="AMUSE 賣得最好嗰支。露水質地薄透到接近透明，但色素企得住——唔會有一般水感唇釉「搽完等於冇搽」嘅問題。",
        bullets=[("露水質感", "極薄極透，唇上冇重量。"),
                 ("純素配方", "唔含動物成分，敏感唇都用得。"),
                 ("薄透可疊", "一層裸唇，三層完整唇妝。"),
                 ("七色可選", "由微風、蜜桃粉到紫水晶。")],
        how="沿唇形塗一層；想飽和啲就等半乾後再疊。")),

    ("amuse-jelfit-tint", "AMUSE Jelfit Tint 果凍唇釉", "唇釉", T_LIP + ", liptint",
     lambda t: "果凍唇釉" in t, dict(
        hook="果凍質地，彈返上嚟。",
        lede="果凍膜貼住唇部，光澤度高但唔黐。食完飯光澤退咗，色仲喺唇上——薄透唔代表唔耐用。",
        bullets=[("果凍彈潤", "水光感強但唔黏。"),
                 ("鎖色唔斑駁", "淡出均勻，唔會淨返唇線。"),
                 ("唔黐頭髮", "風大都唔怕。"),
                 ("六色可選", "由蜜桃、裸粉到焦糖楓葉。")],
        how="沿唇形塗抹；只點唇中央可做漸層唇。")),

    ("amuse-bebe-tint", "AMUSE BEBE Tint 嬰兒感唇釉", "唇釉", T_LIP + ", liptint",
     lambda t: "BEBE 唇釉" in t, dict(
        hook="似天生唇色，唔似搽咗嘢。",
        lede="BEBE 系列全部係低飽和裸調，塗上去唔會覺得「化咗妝」，只會覺得今日唇色好靚——素顏同淡妝最啱。",
        bullets=[("裸調自然", "低飽和度，返工返學都用得。"),
                 ("薄透水潤", "唔會顯唇紋。"),
                 ("純素配方", "唔含動物成分。"),
                 ("五色可選", "由蜜桃、木瓜到無花果同葡萄。")],
        how="沿唇形塗一層，抿唇令顏色均勻。")),

    ("amuse-powder-velvet-tint", "AMUSE Powder Velvet Tint 絲絨霧面唇泥", "唇釉",
     T_LIP + ", liptint", lambda t: "啞光唇泥" in t, dict(
        hook="霧面唇泥，唔會乾到起皮。",
        lede="粉霧質地上唇即刻霧化唇紋，但保留潤度——即係話唔使先搽潤唇膏打底。顯色比 Dew Tint 高一級，適合想要存在感嘅日子。",
        bullets=[("天鵝絨霧感", "柔霧唔乾，唔會起皮。"),
                 ("模糊唇紋", "唇部即刻平滑。"),
                 ("高顯色", "一層就夠飽和。"),
                 ("六色可選", "由稀世蜜桃、柔紗珊瑚到海鹽葡萄。")],
        how="由唇中央向外塗抹，再以指腹拍散邊緣。")),

    ("amuse-powder-lip-cheek", "AMUSE Powder Lip & Cheek 絲絨唇頰膏", "胭脂",
     T_CHEEK + ", blush", lambda t: "啞光唇頰膏" in t, dict(
        hook="唇同頰用同一支，妝感自然統一。",
        lede="唇色同腮紅唔夾，成個妝就會怪。呢支點唇之後順手掃兩下喺頰，色調自動統一——出門帶一支就夠。",
        bullets=[("唇頰兩用", "一支搞掂全臉氣色。"),
                 ("粉霧收尾", "唔油亮亦唔過分啞。"),
                 ("易推開", "喺頰上唔會拉扯底妝。"),
                 ("六色可選", "由蜂蜜、珊瑚泡泡到玫瑰少女。")],
        how="點於唇中央推開；點頰時輕點三下再以指腹拍散。")),

    ("amuse-lip-cheek-healthy-balm", "AMUSE 雙色唇頰兩用膏", "胭脂", T_CHEEK + ", blush",
     lambda t: "雙色唇頰兩用膏" in t, dict(
        hook="一支兩色，深淺自己溝。",
        lede="同一支入面有深淺兩格，可以單用、可以疊、可以溝——即係話同一支喺唔同季節、唔同妝濃度都用得。",
        bullets=[("兩色一支", "深淺自由調配。"),
                 ("唇頰通用", "唇、頰、眼皮都點得。"),
                 ("滋潤配方", "唇上唔會乾。"),
                 ("五款水果調", "桃子、芒果、草莓、葡萄、無花果。")],
        how="淺色鋪底，深色收窄範圍；或兩色一齊沾做自然過渡。")),

    ("amuse-cheek-tok-tok", "AMUSE Cheek Tok Tok 液體腮紅", "胭脂", T_CHEEK + ", blush",
     lambda t: "液體腮紅" in t, dict(
        hook="輕拍兩下，血色由皮膚透出嚟。",
        lede="液體質地一觸即化，指腹輕拍就融入底妝，唔會浮喺粉底上面——望落唔似搽咗嘢。",
        bullets=[("融入底妝", "唔會推花粉底。"),
                 ("免工具", "手指拍兩下就完成。"),
                 ("持久貼服", "口罩擦唔走。"),
                 ("純素配方", "唔含動物成分。")],
        how="指腹沾取米粒大小，點於顴骨後由內向外輕拍。")),

    ("amuse-ceramic-skin-cushion", "AMUSE 白色陶瓷氣墊粉底", "氣墊粉底", T_BASE + ", cushion",
     lambda t: "白色陶瓷氣" in t, dict(
        hook="陶瓷一樣嘅啞緻膚感。",
        lede="出油肌用水光氣墊，中午就會反光到似出汗。陶瓷氣墊做啞緻收尾，控油同時唔會乾——妝面平滑得似上咗釉。",
        bullets=[("啞緻收尾", "控油唔反光。"),
                 ("平滑毛孔", "妝面均勻似陶瓷。"),
                 ("持妝唔氧化", "全日唔變深變黃。"),
                 ("兩個色階", "01 象牙同 1.5 自然色。")],
        how="以粉撲輕拍上臉，由面中央向外推開。")),

    ("amuse-dew-jelly-cushion", "AMUSE 黃色果凍氣墊粉底", "氣墊粉底", T_BASE + ", cushion",
     lambda t: "黃色果凍氣墊" in t, dict(
        hook="果凍質地，薄到似冇上底妝。",
        lede="遮瑕力夠但唔會厚重。果凍粉體貼膚後成一層薄膜，妝感輕透水潤，乾肌用都唔會卡粉。",
        bullets=[("果凍水潤", "貼膚唔卡粉。"),
                 ("薄塗夠遮", "唔使疊厚就均勻。"),
                 ("自然光澤", "亮而唔油。"),
                 ("兩個色階", "01 乾淨同 1.5 自然色。")],
        how="以粉撲輕拍上臉；瑕疵位再輕拍一層。")),

    ("amuse-dew-power-cushion", "AMUSE 粉色防曬氣墊粉底", "氣墊粉底", T_BASE + ", cushion",
     lambda t: "粉色防曬氣墊" in t, dict(
        hook="底妝同防曬，一步做齊。",
        lede="早上趕時間最想慳一步。呢款氣墊自帶防曬，粉調校色同時中和暗黃——搽完唔使再上防曬。",
        bullets=[("自帶防曬", "慳一個步驟。"),
                 ("粉調校色", "中和暗黃，膚色即刻乾淨。"),
                 ("水潤貼膚", "唔會卡粉。"),
                 ("兩個色階", "01 純淨同 1.5 自然色。")],
        how="以粉撲輕拍上臉，由面中央向外推開。")),

    ("amuse-vegan-hand-cream", "AMUSE Vegan Soybean 護手霜", "護手霜", T_BODY + ", hand cream",
     lambda t: "護手霜" in t, dict(
        hook="搽完即刻拎到手機。",
        lede="大豆配方吸收快，唔會留低油膜——搽完可以即刻打字、拎嘢，唔使等。香味柔和唔搶，同香水唔會打架。",
        bullets=[("極速吸收", "唔會黏笠笠。"),
                 ("大豆保濕", "乾燥季節都撐得住。"),
                 ("柔和香味", "唔會蓋過香水。"),
                 ("純素配方", "唔含動物成分。")],
        how="取適量搽於手部，由手背推向指縫。",
        specs=["產地：韓國 Made in Korea", "純素配方 Vegan"])),

    # ── Sanrio collaborations ───────────────────────────────────────
    ("amuse-sanrio-mymelody-cushion", "AMUSE × My Melody 聯名氣墊粉底", "氣墊粉底",
     T_BASE + ", cushion, sanrio, 聯名",
     lambda t: "聯名版" in t and "氣墊粉底" in t and "My Melody" in t, dict(
        hook="My Melody 限定粉盒。",
        lede="AMUSE 皇牌氣墊嘅限定版本——配方一樣，換上 My Melody 嘅粉盒同鏡面設計。限量發售，賣完就冇。",
        bullets=[("限定包裝", "My Melody 粉盒同鏡面。"),
                 ("同正裝配方", "遮瑕同持妝力一樣。"),
                 ("連補充裝", "用完換芯，粉盒留低。"),
                 ("三個色階", "0.5、01 同 1.5，白皮到自然色。")],
        how="以粉撲輕拍上臉，由面中央向外推開。")),

    ("amuse-sanrio-kuromi-cushion", "AMUSE × Kuromi 聯名氣墊粉底", "氣墊粉底",
     T_BASE + ", cushion, sanrio, 聯名",
     lambda t: "聯名版" in t and "氣墊粉底" in t and "Kuromi" in t, dict(
        hook="Kuromi 限定粉盒。",
        lede="AMUSE 皇牌氣墊嘅限定版本——配方一樣，換上 Kuromi 嘅粉盒同鏡面設計。限量發售，賣完就冇。",
        bullets=[("限定包裝", "Kuromi 粉盒同鏡面。"),
                 ("同正裝配方", "遮瑕同持妝力一樣。"),
                 ("連補充裝", "用完換芯，粉盒留低。"),
                 ("三個色階", "0.5、01 同 1.5。")],
        how="以粉撲輕拍上臉，由面中央向外推開。")),

    ("amuse-lip-liner", "AMUSE 唇線筆", "唇線筆", T_LIP + ", lip pencil",
     lambda t: "唇線筆" in t, dict(
        hook="唇形清晰，成個妝就企得住。",
        lede="同一隻唇釉，畫咗唇線同冇畫，望落精神度差好遠。呢支質地夠軟唔會拉扯唇部，可以描邊亦可以填滿全唇。",
        bullets=[("描得準", "筆芯幼細，唇峰唇角都定得住。"),
                 ("唔拉扯唇部", "乾唇都唔會刮。"),
                 ("防走位", "唇釉唔會溢出唇緣。"),
                 ("四色可選", "由陰影色到玫瑰同奶桃。")],
        how="由唇峰向唇角描出輪廓，再填滿全唇或疊唇釉。")),

    ("amuse-sanrio-lip-tint", "AMUSE × Sanrio 聯名唇釉", "唇釉",
     T_LIP + ", liptint, sanrio, 聯名",
     lambda t: "聯名版" in t and ("水光唇釉" in t or "絲絨唇釉" in t), dict(
        hook="限定色，配限定唇釉管。",
        lede="My Melody 同 Kuromi 主題嘅限定色，水光同絲絨兩種質地——管身印住角色圖案，補妝嗰陣都靚。",
        bullets=[("限定色調", "只喺聯名系列出現。"),
                 ("兩種質地", "水光同絲絨，跟妝容揀。"),
                 ("角色管身", "My Melody 同 Kuromi 圖案。"),
                 ("限量發售", "賣完即止。")],
        how="沿唇形塗抹，抿唇令顏色均勻。")),

    ("amuse-sanrio-lip-balm", "AMUSE × Sanrio 聯名 Soda 潤唇膏", "唇膏",
     T_LIP + ", lipbalm, sanrio, 聯名",
     lambda t: "聯名版" in t and "潤唇膏" in t, dict(
        hook="汽水感潤唇膏，帶少少色。",
        lede="潤唇為主、帶色為輔——唔想化妝但又想唇色好啲嗰啲日子最啱。管身係 My Melody 同 Kuromi 限定設計。",
        bullets=[("滋潤為主", "唔會乾，唔會起皮。"),
                 ("微微帶色", "自然唇色，唔似化咗妝。"),
                 ("汽水感光澤", "清爽唔黐。"),
                 ("限定包裝", "兩款角色設計。")],
        how="直接塗於唇部，隨時補搽。")),

    ("amuse-sanrio-cheek-balm", "AMUSE × Sanrio 聯名胭脂膏（夾子款）", "胭脂",
     T_CHEEK + ", blush, sanrio, 聯名",
     lambda t: "聯名版" in t and "胭脂膏" in t, dict(
        hook="夾子造型，掛喺袋上都得。",
        lede="膏狀胭脂，指腹輕拍就融入底妝。外殼做成夾子造型，可以夾喺袋上或者掛繩上——補妝隨手可及。",
        bullets=[("夾子外殼", "掛喺袋上，補妝方便。"),
                 ("膏狀貼膚", "融入底妝，唔會浮。"),
                 ("免工具", "手指拍兩下就完成。"),
                 ("限定角色", "My Melody 同 Kuromi 兩款。")],
        how="指腹沾取少量，點於顴骨後向外輕拍。")),

    ("amuse-sanrio-eye-palette", "AMUSE × Sanrio 聯名眼影盤", "眼影",
     T_EYE + ", eyeshadow, palette, sanrio, 聯名",
     lambda t: "聯名版" in t and "眼影盤" in t, dict(
        hook="限定盤，配色跟住角色走。",
        lede="My Melody 嘅粉調同 Kuromi 嘅紫黑調，各自一盤。啞光同閃片齊全，順住格數用就係完整眼妝。",
        bullets=[("角色配色", "粉調同紫黑調各一盤。"),
                 ("啞光＋閃片", "質地齊全。"),
                 ("粉質細滑", "同正裝同一配方。"),
                 ("限量發售", "賣完即止。")],
        how="淺色打底，中間色暈染褶位，深色壓眼尾，閃片點眼中央。")),
]


def load_shades():
    wb = openpyxl.load_workbook(STOCK_XLSX)
    ws = wb.active
    h = [c.value for c in ws[1]]
    iT, iV, iQ, iP, iB = (h.index(x) for x in
                          ("Title", "Vendor", "Variant Inventory Qty",
                           "Variant Price", "Variant Barcode"))
    return [(str(r[iT]).strip(), str(r[iB]).strip(), r[iP], r[iQ] or 0)
            for r in ws.iter_rows(min_row=2, values_only=True)
            if r[iV] and "amuse" in str(r[iV]).lower()]


def shade_name(title):
    """The shade is whatever the supplier put in brackets or after a #.
    Collab rows read "… My Melody 氣墊粉底 0.5#", where the bare number
    would collide with the Kuromi row of the same shade."""
    m = re.search(r"(My Melody|Kuromi)[^\d]*([\d.]+)#", title)
    if m:
        return f"{m.group(1)} {m.group(2)}"
    m = re.search(r"[（(]\s*(.+?)\s*[）)]", title)
    if m:
        return m.group(1)
    m = re.search(r"#\s*(\d+[^#]*)$", title)
    if m:
        return m.group(1).strip()
    m = re.search(r"(\d+[.\d]*\s*\w[\w ]*)$", title)
    return m.group(1).strip() if m else title


P = {}
rows = load_shades()
used = set()
for slug, title, ptype, tags, match, copy in LINES:
    picked = [(shade_name(t), b, q, p) for t, b, p, q in rows
              if match(t) and b not in used]
    used.update(b for _, b, _, _ in picked)
    if not picked:
        continue
    P[slug] = dict(title=title, type=ptype, tags=tags,
                   price=max(p for _, _, _, p in picked) or 0,
                   specs=copy.get("specs", SPECS),
                   hook=copy["hook"], lede=copy["lede"],
                   bullets=copy["bullets"], how=copy["how"],
                   shades=sorted((n, b, q) for n, b, q, _ in picked))

_missed = [t for t, b, _, _ in rows if b not in used]
if _missed:
    print(f"未分組（{len(_missed)}）: " + "; ".join(_missed[:6]))

run(__name__, VENDOR, P, "amuse")
