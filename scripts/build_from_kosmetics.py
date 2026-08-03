#!/usr/bin/env python3
"""
Build a brand from kosmetics.com.hk, the HK/Macau distributor.

Best source we have found. It runs on Shopify, so products.json hands
over Traditional Chinese titles, descriptions, shade names and imagery as
structured data — no scraping, and because the images already sit on
cdn.shopify.com our own store can fetch them straight from there. Nothing
gets mirrored into this repo for these brands.

SKUs are matched to ours by BARCODE, not by name. Our supplier list has
demonstrable naming errors (a CLIO brow pencil filed as a mascara), and
the barcode is what is printed on the box that ships.

    python3 scripts/build_from_kosmetics.py peripera --dry-run
    python3 scripts/build_from_kosmetics.py peripera
"""
import argparse
import collections
import json
import os
import re
import sys
import urllib.request

import openpyxl

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from publish import publish  # noqa: E402

DIST = "https://kosmetics.com.hk"
UA = ("Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 "
      "(KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36")
STOCK_XLSX = "/Volumes/core/下載/cosmetic_products_shopify.xlsx"

# distributor product_type -> (our productType, section tag)
TYPE_MAP = {
    "Face": ("胭脂", "修容"),
    "Lip": ("唇釉", "唇妝"),
    "Eye": ("眼影", "眼妝"),
    "Base": ("底妝", "底妝"),
}


def distributor_products(vendor):
    req = urllib.request.Request(
        f"{DIST}/collections/{vendor}/products.json?limit=250", headers={"User-Agent": UA})
    return json.load(urllib.request.urlopen(req, timeout=60))["products"]


def our_rows(vendor):
    wb = openpyxl.load_workbook(STOCK_XLSX)
    ws = wb.active
    h = [c.value for c in ws[1]]
    iT, iV, iQ, iP, iB = (h.index(x) for x in
                          ("Title", "Vendor", "Variant Inventory Qty",
                           "Variant Price", "Variant Barcode"))
    out = []
    for r in ws.iter_rows(min_row=2, values_only=True):
        if r[iV] and vendor.lower() in str(r[iV]).lower():
            out.append({"title": str(r[iT]), "barcode": str(r[iB]).strip(),
                        "price": r[iP], "qty": r[iQ] or 0})
    return out


def clean_body(html, title):
    """Their body is one flat block of ✧ lines plus a shade rundown. Keep
    the selling points; drop the shade list, which is their line-up and
    not our stock."""
    text = re.sub(r"<[^>]+>", "\n", html or "")
    lines = [ln.strip() for ln in text.split("\n") if ln.strip()]
    keep = []
    for ln in lines:
        if ln == title or ln.startswith(("🌹", "🍑", "🍫")) and re.match(r"^.{0,3}\s*\d+\s", ln):
            continue
        if re.match(r"^[\d]{1,3}\s", ln):      # "28 浪漫玫瑰 Romantic Rose - ..."
            continue
        keep.append(ln)
    return keep


def description(prod, keep):
    h = []
    ticks = [k.lstrip("✧✓· ").strip() for k in keep if k.startswith(("✧", "✓"))]
    rest = [k for k in keep if not k.startswith(("✧", "✓"))]
    for r in rest[:2]:
        h.append(f"<p><strong>{r}</strong></p>" if r == rest[0] else f"<p>{r}</p>")
    if ticks:
        h.append("<ul>" + "".join(f"<li>{t}</li>" for t in ticks) + "</ul>")
    h.append("<ul><li>產地：韓國 Made in Korea</li></ul>")
    imgs = [i["src"] for i in prod["images"]]
    if len(imgs) > 1:
        h.append('<div class="product-detail-images">'
                 + "".join(f'<img src="{u}" alt="{prod["title"]} 產品介紹" loading="lazy">'
                           for u in imgs[1:])
                 + "</div>")
    return "".join(h)


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("vendor")
    ap.add_argument("--dry-run", action="store_true")
    args = ap.parse_args()

    dist = distributor_products(args.vendor)
    index = {}
    for p in dist:
        for v in p["variants"]:
            for key in (v.get("sku"), v.get("barcode")):
                if key:
                    index[str(key).strip()] = (p, v)

    groups, unmatched = collections.OrderedDict(), []
    for row in our_rows(args.vendor):
        hit = index.get(row["barcode"])
        if not hit:
            unmatched.append(row)
            continue
        p, v = hit
        g = groups.setdefault(p["handle"], {"prod": p, "shades": []})
        g["shades"].append({
            "name": v["title"],
            "barcode": row["barcode"],
            "qty": row["qty"],
            "price": row["price"],
            "image": (v.get("featured_image") or {}).get("src"),
        })

    vendor_label = dist[0]["vendor"] if dist else args.vendor
    for handle, g in groups.items():
        p = g["prod"]
        ptype, section = TYPE_MAP.get(p["product_type"], ("彩妝", "彩妝"))
        imgs = [i["src"] for i in p["images"]]
        # keep every variant swatch reachable as a gallery image
        imgs += [s["image"] for s in g["shades"] if s.get("image")]
        shades = sorted(g["shades"], key=lambda s: s["name"])
        price = max(s["price"] for s in shades)
        item = {
            "handle": handle,
            "title": p["title"],
            "descriptionHtml": description(p, clean_body(p.get("body_html"), p["title"])),
            "vendor": vendor_label,
            "productType": ptype,
            "tags": [vendor_label, "K-Beauty", "彩妝", section, "makeup"],
            "status": "ACTIVE",
            "option_name": "色號",
            "price": price,
            "images": list(dict.fromkeys(imgs)),
            "shades": shades,
        }
        print(f'{len(shades):>2} 色  {len(item["images"]):>2} 圖  ${price:<5} {p["title"]}')
        if not args.dry_run:
            r = publish(item)
            print(f"        -> {r['handle']}  {r['variants']} variants, "
                  f"{r['media']} media, {r['channels']} channels")

    if unmatched:
        print(f"\n條碼對唔到經銷商目錄（{len(unmatched)} 個 SKU）：")
        for r in unmatched:
            print(f"    {r['barcode']}  ${r['price']} 庫存 {r['qty']}  {r['title']}")


if __name__ == "__main__":
    main()
