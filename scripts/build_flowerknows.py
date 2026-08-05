#!/usr/bin/env python3
"""
Build and publish the Flower Knows range.

A Chinese brand, so our supplier titles are already Traditional Chinese
and only imagery is wanted. Its global store is Shopify, and because its
images sit on cdn.shopify.com our own store fetches them directly —
nothing is mirrored.

The store carries no barcodes, so lines join by product name. The join is
checked by variant count: every one of the nine matched lines has the
same number of shades on both sides, which is what makes the pairing
trustworthy rather than a guess.

    python3 scripts/build_flowerknows.py [--dry-run]
"""
import argparse
import json
import os
import re
import sys
import urllib.request

import openpyxl

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from publish import publish  # noqa: E402

STORE = "https://flowerknows.co"
UA = ("Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 "
      "(KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36")
STOCK_XLSX = "/Volumes/core/下載/cosmetic_products_shopify.xlsx"
VENDOR = "花知曉 Flower Knows"

T_LIP = "花知曉, Flower Knows, 彩妝, 唇妝, makeup, lip"
T_CHEEK = "花知曉, Flower Knows, 彩妝, 修容, makeup, cheek"
T_EYE = "花知曉, Flower Knows, 彩妝, 眼妝, makeup, eye"
T_BASE = "花知曉, Flower Knows, 彩妝, 底妝, makeup, base"
T_LIFE = "花知曉, Flower Knows, 生活風格, lifestyle"
SPECS = ["甜心小熊系列"]

# our title keyword -> (store title, our title, type, tags, copy)
LINES = [
    ("成膜唇凍", "The Sweetie Bear Coating Lip Jelly",
     "花知曉 甜心小熊 成膜唇凍", "唇釉", T_LIP + ", liptint", dict(
        hook="果凍質地，成膜之後唔甩色。",
        lede="唇凍上唇薄透水潤，等佢成膜之後色素就鎖住——飲嘢食飯都唔會印落杯邊。甜心小熊系列嘅包裝本身已經值得擺出嚟。",
        bullets=[("果凍水潤", "薄透唔黐笠。"), ("成膜鎖色", "唔印杯唔甩色。"),
                 ("九色最闊", "全線色域最齊。"), ("收藏級包裝", "小熊造型，擺枱都好睇。")],
        how="沿唇形塗一層，靜待十數秒成膜。")),

    ("絲緞腮紅", "The Sweetie Bear Silk Satin Blush",
     "花知曉 甜心小熊 絲緞腮紅", "胭脂", T_CHEEK + ", blush", dict(
        hook="絲緞光澤，唔係粉感。",
        lede="粉體極細，掃上面會有一層柔和光澤而唔係啞粉——所以唔會令面部顯乾，反而好似皮膚本身有光。",
        bullets=[("絲緞質感", "柔和光澤，唔會粉感。"), ("易控色", "逐層疊到啱為止。"),
                 ("貼膚持久", "唔飛粉。"), ("五色可選", "由蜜桃到莓調。")],
        how="以胭脂掃沾取，由笑肌向太陽穴輕掃。")),

    ("六色眼影", "The Sweetie Bear 6-Color Makeup Palette",
     "花知曉 甜心小熊 六色眼影盤", "眼影", T_EYE + ", eyeshadow, palette", dict(
        hook="六格，配色已經幫你諗好。",
        lede="啞光同閃片齊全，由淺到深排好，順住用就係完整眼妝。盤面同外殼都係小熊主題——新手友好，收藏都值。",
        bullets=[("配色已諗好", "順住格數用就得。"), ("啞光＋閃片", "質地齊全。"),
                 ("粉質細滑", "唔飛粉。"), ("收藏級包裝", "小熊浮雕盤面。")],
        how="淺色打底，中間色暈染褶位，深色壓眼尾，閃片點眼中央。")),

    ("雙頭染眉膏", "The Sweetie Bear Dual-Ended Brow Gel & Pencil",
     "花知曉 甜心小熊 雙頭染眉膏／眉筆", "眉筆", T_EYE + ", brow", dict(
        hook="一支兩頭：畫同染。",
        lede="一頭幼細眉筆補稀疏位，另一頭染眉膏調整眉色同定型——染完頭髮之後眉色唔夾，用呢支就搞掂。",
        bullets=[("一支兩用", "眉筆＋染眉膏。"), ("防水持久", "全日唔甩色。"),
                 ("柔和眉色", "自然貼近髮色。"), ("三色可選", "跟髮色揀。")],
        how="先以眉筆補稀疏位，再用染眉膏順住毛流刷勻。")),

    ("四色遮瑕盤", "The Sweetie Bear 4-Color Concealer Palette",
     "花知曉 甜心小熊 四色遮瑕盤", "底妝", T_BASE + ", concealer", dict(
        hook="四格校色，唔同瑕疵用唔同色。",
        lede="泛紅要綠調、黑眼圈要蜜桃調——一盤四格可以按位置調配，唔使買幾支。體積細，帶出街補妝都方便。",
        bullets=[("四格校色", "按瑕疵類型調配。"), ("貼膚唔厚重", "唔會有面具感。"),
                 ("旅行啱用", "一盤頂幾支。"), ("兩款可選", "跟膚色深淺揀。")],
        how="以遮瑕刷沾取對應色，點於瑕疵位輕拍推開。")),

    ("迷你蜜粉餅", "The Sweetie Bear Mini Setting Powder",
     "花知曉 甜心小熊 迷你蜜粉餅", "底妝", T_BASE + ", powder", dict(
        hook="細到放得入銀包。",
        lede="粉體幼細，補妝時輕按 T 字位就控到油，唔會愈補愈厚。細細個帶出街唔佔位——小熊粉盒拎出嚟補妝都好睇。",
        bullets=[("極細粉體", "唔會浮粉見紋。"), ("控油定妝", "補妝唔會厚。"),
                 ("迷你尺寸", "銀包都放得落。"), ("收藏級包裝", "小熊粉盒。")],
        how="以粉撲沾取，輕按於 T 字位或出油位。")),

    ("香水", "The Sweetie Bear Perfume",
     "花知曉 甜心小熊 香水", "香水", "花知曉, Flower Knows, 香氛, fragrance", dict(
        hook="甜而唔膩嘅小熊香。",
        lede="前調果甜、中後段轉成柔和奶香同木質，所以唔會似一般甜香咁快膩。瓶身本身就係擺設。",
        bullets=[("層次分明", "由果甜轉奶香木質。"), ("甜而唔膩", "日常戴得住。"),
                 ("留香溫和", "唔會搶過人。"), ("收藏級瓶身", "小熊造型。")],
        how="噴於手腕、頸側或耳後脈搏位置。")),

    ("手持鏡", "The Sweetie Bear Hand Mirror",
     "花知曉 甜心小熊 手持鏡", "配件", T_LIFE + ", mirror", dict(
        hook="化妝袋入面最好睇嗰件。",
        lede="小熊造型手持鏡，鏡面夠大照到全臉。實用之餘更加係擺設——放喺枱面或者化妝袋都吸睛。",
        bullets=[("小熊造型", "系列同款設計。"), ("鏡面夠大", "照到全臉。"),
                 ("手感紮實", "唔似平價塑膠。"), ("三款可選", "配色同系列一致。")],
        how="隨身攜帶或放於梳妝枱。")),

    ("點彩刷", "The Sweetie Bear Rounded Blush Brush",
     "花知曉 甜心小熊 點彩刷", "化妝工具", T_LIFE + ", brush", dict(
        hook="圓頭刷，胭脂唔會落錯位。",
        lede="圓頭設計令粉量分佈均勻，掃上面唔會有明顯邊界。刷毛柔軟唔拮面，握柄同系列一樣係小熊主題。",
        bullets=[("圓頭設計", "上色均勻無邊界。"), ("刷毛柔軟", "唔拮面。"),
                 ("易清洗", "唔易掉毛。"), ("系列同款", "小熊握柄。")],
        how="沾取胭脂後喺手背拍走多餘粉量，再由笑肌向外輕掃。")),
]


def store_products():
    req = urllib.request.Request(f"{STORE}/products.json?limit=250",
                                 headers={"User-Agent": UA})
    return json.load(urllib.request.urlopen(req, timeout=60))["products"]


def our_rows():
    wb = openpyxl.load_workbook(STOCK_XLSX)
    ws = wb.active
    h = [c.value for c in ws[1]]
    iT, iV, iQ, iP, iB = (h.index(x) for x in
                          ("Title", "Vendor", "Variant Inventory Qty",
                           "Variant Price", "Variant Barcode"))
    return [(str(r[iT]).strip(), str(r[iB]).strip(), r[iP], r[iQ] or 0)
            for r in ws.iter_rows(min_row=2, values_only=True)
            if r[iV] and ("花知" in str(r[iV]) or "flower" in str(r[iV]).lower())]


def shade_of(title, keyword):
    tail = title.split(keyword, 1)[-1].strip(" （）()[]#")
    return re.sub(r"\s{2,}", " ", tail).strip() or "單一規格"


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--dry-run", action="store_true")
    args = ap.parse_args()

    store = {p["title"]: p for p in store_products()}
    rows = our_rows()
    used = set()

    for keyword, store_title, title, ptype, tags, copy in LINES:
        ours = [(shade_of(t, keyword), b, q, p) for t, b, p, q in rows
                if keyword in t and b not in used]
        src = store.get(store_title)
        if not ours or not src:
            print(f"  ?? {title}: ours={len(ours)} store={'y' if src else 'n'}")
            continue
        if len(ours) != len(src["variants"]):
            print(f"  !! {title}: 我哋 {len(ours)} 色 vs 官方 {len(src['variants'])} 色，跳過")
            continue
        used.update(b for _, b, _, _ in ours)

        imgs = [i["src"] for i in src["images"]]
        body = ("".join(f'<p><strong>{copy["hook"]}</strong></p><p>{copy["lede"]}</p>')
                + "<ul>" + "".join(f"<li><strong>{t}</strong>——{d}</li>"
                                   for t, d in copy["bullets"]) + "</ul>"
                + f'<p><strong>用法</strong><br>{copy["how"]}</p>'
                + "<ul>" + "".join(f"<li>{s}</li>" for s in SPECS) + "</ul>")
        if len(imgs) > 1:
            body += ('<div class="product-detail-images">'
                     + "".join(f'<img src="{u}" alt="{title} 產品介紹" loading="lazy">'
                               for u in imgs[1:])
                     + "</div>")

        item = {
            "handle": "flowerknows-" + re.sub(r"[^a-z0-9]+", "-",
                                              store_title.lower()).strip("-"),
            "title": title, "descriptionHtml": body, "vendor": VENDOR,
            "productType": ptype, "tags": [x.strip() for x in tags.split(",")],
            "status": "ACTIVE", "option_name": "款式",
            "price": max(p for _, _, _, p in ours), "images": imgs,
            "shades": [{"name": n, "barcode": b, "qty": q} for n, b, q, _ in ours],
        }
        print(f'{len(ours):>2} 色  {len(imgs):>2} 圖  {title}')
        if not args.dry_run:
            r = publish(item)
            print(f"        -> {r['handle']}  {r['variants']} variants, "
                  f"{r['media']} media, {r['channels']} channels")

    left = [t for t, b, _, _ in rows if b not in used]
    if left:
        print(f"\n官方店冇對應（{len(left)}）：" + "; ".join(left[:6]))


if __name__ == "__main__":
    main()
