#!/usr/bin/env python3
"""
Build and publish the TIR TIR range.

TIR TIR is a base-makeup brand: six full-size cushions and four minis
that share the Mask Fit name but are different formulas in different
cases, so each is its own product rather than a shade of one. It is also
the first brand here carrying real skincare, which is tagged 護膚 so it
lands on the skincare page instead of among the makeup.

Shades come from the workbook at build time, never transcribed.

    python3 scripts/build_tirtir.py mirror
    python3 scripts/build_tirtir.py publish [--dry-run]
"""
import re

import openpyxl

from brand_build import run

VENDOR = "TIRTIR"
STOCK_XLSX = "/Volumes/core/下載/cosmetic_products_shopify.xlsx"

T_BASE = "TIRTIR, K-Beauty, 彩妝, 底妝, makeup, base"
T_LIP = "TIRTIR, K-Beauty, 彩妝, 唇妝, makeup, lip"
T_SKIN = "TIRTIR, K-Beauty, 護膚, skincare"
SPECS = ["產地：韓國 Made in Korea"]

CUSHION = "以粉撲輕拍上臉，由面中央向外推開；瑕疵位再輕拍一層。"

# slug -> (title, type, tags, title matcher, copy)
LINES = [
    ("tirtir-mask-fit-red-cushion", "TIRTIR Mask Fit 紅色氣墊粉底", "氣墊粉底",
     T_BASE + ", cushion", lambda t: "RED CUSHION" in t.upper(), dict(
        hook="全球賣爆嗰個紅盒。",
        lede="TIR TIR 一戰成名嘅就係呢隻。遮瑕力高但唔會厚重，妝感貼服自然，而且色階齊到深膚色都搵到位——韓國以外賣得最好嘅氣墊之一。",
        bullets=[("高遮瑕", "痘印同泛紅一層搞掂。"),
                 ("貼服唔厚重", "唔會有面具感。"),
                 ("持妝一整日", "唔氧化唔脫妝。"),
                 ("連補充裝", "用完換芯。")], how=CUSHION)),

    ("tirtir-mask-fit-pink-cushion", "TIRTIR Mask Fit All Cover 粉色氣墊粉底", "氣墊粉底",
     T_BASE + ", cushion", lambda t: "ALL COVER CUSHION(PINK)" in t.upper(), dict(
        hook="紅盒嘅水潤版本。",
        lede="同樣係全覆蓋遮瑕，但收尾水潤啲、光澤自然啲。乾肌用紅盒會覺得緊，粉盒就啱——遮瑕力一樣，膚感唔同。",
        bullets=[("全覆蓋遮瑕", "同紅盒同級遮瑕力。"),
                 ("水潤收尾", "乾肌唔會卡粉。"),
                 ("自然光澤", "亮而唔油。"),
                 ("連補充裝", "用完換芯。")], how=CUSHION)),

    ("tirtir-mask-fit-ai-filter-cushion", "TIRTIR Mask Fit AI Filter 氣墊粉底", "氣墊粉底",
     T_BASE + ", cushion", lambda t: "AI FILTER" in t.upper() and "MINI" not in t.upper(), dict(
        hook="似開咗濾鏡，但係真皮膚。",
        lede="橙盒嘅賣點係「柔焦」——粉體會散射光線，令毛孔同細紋淡化，所以望落平滑但唔會覺得搽咗好厚粉。",
        bullets=[("柔焦毛孔", "妝面平滑均勻。"),
                 ("輕薄貼膚", "薄塗已經夠勻。"),
                 ("持妝唔氧化", "全日唔變深變黃。"),
                 ("連補充裝", "用完換芯。")], how=CUSHION)),

    ("tirtir-mask-fit-aura-cushion", "TIRTIR Mask Fit Aura 銀色氣墊粉底", "氣墊粉底",
     T_BASE + ", cushion", lambda t: "AURA CUSHION" in t.upper(), dict(
        hook="由內透出嚟嘅光。",
        lede="銀盒行嘅係水光路線——唔係表面油亮，而係皮膚本身透出光感。想要韓系「玻璃肌」嗰種妝感就揀呢隻。",
        bullets=[("水光妝感", "光由內透出，唔係表面反光。"),
                 ("薄透貼膚", "唔會蓋住膚質。"),
                 ("保濕配方", "乾肌全日唔緊繃。"),
                 ("連補充裝", "用完換芯。")], how=CUSHION)),

    ("tirtir-mask-fit-crystal-mesh-cushion", "TIRTIR Mask Fit Crystal Mesh 氣墊粉底",
     "氣墊粉底", T_BASE + ", cushion", lambda t: "CRYSTAL MESH" in t.upper(), dict(
        hook="網面出粉，唔會一撳出一堆。",
        lede="Crystal Mesh 網面令每次出粉量剛好，所以薄薄一層就夠勻，唔會浪費亦唔會搽厚。妝感通透，適合想要輕妝嘅日子。",
        bullets=[("網面控量", "出粉剛好，唔會過量。"),
                 ("通透妝感", "輕薄但均勻。"),
                 ("貼膚持久", "唔卡粉唔脫妝。"),
                 ("三個色階", "白皮到自然色。")], how=CUSHION)),

    ("tirtir-mask-fit-cool-sun-cushion", "TIRTIR Mask Fit Cool Blue 防曬氣墊", "氣墊粉底",
     T_BASE + ", cushion, suncare", lambda t: "SUN CUSHION" in t.upper(), dict(
        hook="防曬做成氣墊，補得到。",
        lede="防曬最大問題係補唔到——化咗妝之後總唔會再搽一層乳霜。做成氣墊就解決咗：輕拍兩下就補到，唔會推花底妝。",
        bullets=[("可補搽防曬", "唔會推花底妝。"),
                 ("清涼膚感", "夏天用唔會焗。"),
                 ("輕薄不泛白", "唔會有防曬嘅白膜。"),
                 ("底妝上可直接用", "定妝同防曬一次過。")],
        how="化妝後以粉撲輕拍需要補防曬嘅位置。")),

    ("tirtir-mask-fit-red-cushion-mini", "TIRTIR Mask Fit 紅色氣墊粉底（迷你）",
     "氣墊粉底", T_BASE + ", cushion", lambda t: "RED MINI CUSHION" in t.upper(), dict(
        hook="紅盒嘅隨身版。",
        lede="同正裝一樣嘅配方同遮瑕力，體積細一半——袋住補妝唔佔位，或者想先試色都啱。",
        bullets=[("同正裝配方", "遮瑕同持妝力一樣。"),
                 ("隨身尺寸", "化妝袋唔佔位。"),
                 ("試色化算", "想試新色階唔使買正裝。"),
                 ("三個色階", "白皮到自然色。")], how=CUSHION)),

    ("tirtir-mask-fit-ai-filter-cushion-mini", "TIRTIR Mask Fit AI Filter 氣墊粉底（迷你）",
     "氣墊粉底", T_BASE + ", cushion", lambda t: "AI FILTER" in t.upper() and "MINI" in t.upper(),
     dict(hook="柔焦橙盒嘅隨身版。",
          lede="同正裝一樣嘅柔焦配方，體積細一半。中午補妝或者旅行帶出去都方便。",
          bullets=[("同正裝配方", "柔焦效果一樣。"),
                   ("隨身尺寸", "袋住補妝。"),
                   ("試色化算", "唔使買正裝就試到。"),
                   ("三個色階", "白皮到自然色。")], how=CUSHION)),

    ("tirtir-mask-fit-pink-cushion-mini", "TIRTIR Mask Fit All Cover 粉色氣墊粉底（迷你）",
     "氣墊粉底", T_BASE + ", cushion",
     lambda t: "PINK CUSHION MINI" in t.upper() or "ALL COVER PINK CUSHION MINI" in t.upper(), dict(
        hook="粉盒嘅隨身版。",
        lede="水潤全覆蓋，體積細一半。乾肌中午補妝最啱——唔會愈補愈厚。",
        bullets=[("同正裝配方", "遮瑕同水潤度一樣。"),
                 ("隨身尺寸", "袋住補妝。"),
                 ("乾肌友好", "補妝唔會卡粉。"),
                 ("三個色階", "白皮到自然色。")], how=CUSHION)),

    ("tirtir-mask-fit-aura-cushion-mini", "TIRTIR Mask Fit Aura 銀色氣墊粉底（迷你）",
     "氣墊粉底", T_BASE + ", cushion",
     lambda t: "AURA SILVER CUSHION MINI" in t.upper(), dict(
        hook="水光銀盒嘅隨身版。",
        lede="同正裝一樣嘅水光配方，體積細一半。想全日保持玻璃肌感就袋住佢。",
        bullets=[("同正裝配方", "水光效果一樣。"),
                 ("隨身尺寸", "袋住補妝。"),
                 ("保濕唔緊繃", "補妝唔會乾。"),
                 ("兩個色階", "象牙同自然色。")], how=CUSHION)),

    ("tirtir-mask-fit-red-foundation", "TIRTIR Mask Fit Red 粉底液", "底妝",
     T_BASE + ", foundation", lambda t: "RED FOUNDATION" in t.upper(), dict(
        hook="紅盒氣墊嘅粉底液版本。",
        lede="想要氣墊嘅遮瑕力但鍾意用刷或者海綿上妝——呢支就係。同樣高遮瑕、同樣貼服，但可以自己控制厚薄。",
        bullets=[("高遮瑕", "同紅盒氣墊同級。"),
                 ("厚薄自控", "用刷或海綿自己調。"),
                 ("持妝一整日", "唔氧化唔脫妝。"),
                 ("三個色階", "白皮到自然色。")],
        how="以粉底刷或海綿由面中央向外推開。")),

    ("tirtir-glide-hide-blurring-concealer", "TIRTIR Glide & Hide 柔焦遮瑕", "底妝",
     T_BASE + ", concealer", lambda t: "Concealer" in t, dict(
        hook="遮到，但唔會卡紋。",
        lede="眼下遮瑕最怕落粉之後全部卡喺細紋度。呢支柔焦質地貼膚唔積聚，笑起上嚟都唔會裂，而且色階多到深膚色都揀到。",
        bullets=[("唔卡細紋", "眼下用一整日都平滑。"),
                 ("柔焦質地", "遮蓋之餘淡化紋理。"),
                 ("易推開", "唔會推花底妝。"),
                 ("色階齊全", "由 0.5N 到 8W。")],
        how="點於瑕疵位，以指腹或遮瑕刷由中央向外輕拍。")),

    ("tirtir-mask-fit-makeup-fixer", "TIRTIR Mask Fit 定妝噴霧", "底妝",
     T_BASE + ", setting spray", lambda t: "MAKE UP FIXER" in t.upper(), dict(
        hook="戴口罩都唔會甩妝。",
        lede="Mask Fit 個名唔係講笑——成膜之後底妝唔會轉印到口罩上。噴完唔會有黏笠感，亦唔會令妝面變啞。",
        bullets=[("防轉印", "口罩同衣領都唔會沾。"),
                 ("唔黏笠", "噴完即刻乾。"),
                 ("唔改變妝感", "唔會變啞或變油。"),
                 ("80ml 大容量", "夠用好耐。")],
        how="化妝後距離面部 20cm 均勻噴灑，待乾。")),

    ("tirtir-mask-fit-makeup-cool-fixer", "TIRTIR Mask Fit 清涼定妝噴霧", "底妝",
     T_BASE + ", setting spray", lambda t: "COOL FIXER" in t.upper(), dict(
        hook="定妝，順便降溫。",
        lede="夏天出汗最容易溶妝。清涼版本噴上面即刻降溫收毛孔，同時定住底妝——出街前同中午補妝都用得。",
        bullets=[("即時清涼", "降溫收毛孔。"),
                 ("防轉印", "口罩唔會沾妝。"),
                 ("控油定妝", "出汗都唔溶妝。"),
                 ("80ml 大容量", "夠用好耐。")],
        how="化妝後或中途補妝時距離 20cm 均勻噴灑。")),

    ("tirtir-waterism-glow-tint", "TIRTIR Waterism Glow 水光唇釉", "唇釉",
     T_LIP + ", liptint", lambda t: "Waterism Glow Tint" in t, dict(
        hook="水光但唔黐。",
        lede="唇釉要夠亮通常代表好黐。呢支上唇薄透如水膜，光澤度高但唔會黐頭髮，飲水食嘢之後色仲喺度。",
        bullets=[("水光唔黐", "風大都唔黐面。"),
                 ("持色自然", "淡出均勻唔斑駁。"),
                 ("含護唇成分", "唔會愈搽愈乾。"),
                 ("四色可選", "玫瑰、珊瑚、無花果、蘇格蘭調。")],
        how="沿唇形塗抹，抿唇令顏色均勻。")),

    ("tirtir-waterism-glow-melting-balm", "TIRTIR Waterism 煥彩融化潤唇膏", "唇膏",
     T_LIP + ", lipbalm", lambda t: "MELTING BALM" in t.upper() or "潤唇膏" in t or "潤澤唇膏" in t,
     dict(hook="一觸即融，潤到唔使補。",
          lede="膏體遇到體溫即刻融開，滑過唇部就化成一層水潤薄膜。色淡到似天生唇色，返工返學隨手搽兩下就得。",
          bullets=[("入口即融質地", "唔使用力，唔會拉扯唇部。"),
                   ("高滋潤", "乾唇都唔會起皮。"),
                   ("微微帶色", "自然唇色，唔似化咗妝。"),
                   ("三色可選", "玫瑰、珊瑚、沙調。")],
          how="直接塗於唇部，隨時補搽。")),

    # ── skincare, tagged 護膚 so it lands on the skincare page ──────
    ("tirtir-ceramic-cream", "TIRTIR 保濕陶瓷面霜", "面霜", T_SKIN + ", moisturizer",
     lambda t: "陶瓷面霜" in t, dict(
        hook="搽完摸落似陶瓷。",
        lede="質地綿密但唔笠，吸收之後皮膚表面平滑到似上咗釉。屏障受損、乾到起皮嗰陣用，第二朝會見到分別。",
        bullets=[("高保濕", "乾燥季節都撐得住。"),
                 ("修護屏障", "泛紅同脫皮都改善。"),
                 ("唔笠唔焗", "上妝前用都得。"),
                 ("50ml 大容量", "全臉頸部都夠用。")],
        how="潔面爽膚後取適量塗於全臉，由內向外輕輕按壓至吸收。")),

    ("tirtir-ceramic-milk-ampoule", "TIRTIR 陶瓷牛奶安瓶", "精華", T_SKIN + ", ampoule, serum",
     lambda t: "Ceramic Milk Ampoule" in t or "陶瓷牛奶安瓶" in t, dict(
        hook="牛奶質地，吸收快到唔覺。",
        lede="安瓶通常濃到黐，呢支反而似牛奶——輕薄好吸收，但保濕同修護成分濃度唔低。想加強保養又怕悶痘就用呢種。",
        bullets=[("極速吸收", "唔會黐笠，後續保養照上。"),
                 ("集中保濕", "乾燥同緊繃即刻紓緩。"),
                 ("修護膚質", "長期用改善粗糙。"),
                 ("敏感肌可用", "配方溫和。")],
        how="爽膚後取 2–3 滴於掌心，按壓於全臉；乾燥位可疊多一層。")),

    ("tirtir-milk-skin-toner", "TIRTIR 牛奶滋養柔潤爽膚水（輕盈版）", "爽膚水",
     T_SKIN + ", toner", lambda t: "爽膚水" in t, dict(
        hook="爽膚水，但有乳液嘅滋潤度。",
        lede="輕盈版質地水感，但保濕力似乳液——即係話夏天可以單搽呢一支，唔使再疊面霜。用化妝棉濕敷仲可以當急救面膜。",
        bullets=[("水感高保濕", "夏天可單用。"),
                 ("即時軟化角質", "後續保養更易吸收。"),
                 ("可濕敷", "當急救面膜用。"),
                 ("150ml 大容量", "全身用都夠。")],
        how="潔面後以化妝棉或掌心輕拍全臉；乾燥時可濕敷 5 分鐘。")),

    ("tirtir-collagen-eye-cream", "TIRTIR 膠原蛋白緊緻眼霜", "眼霜", T_SKIN + ", eye cream",
     lambda t: "眼霜" in t, dict(
        hook="眼下細紋，由保濕開始。",
        lede="眼周乾就會顯紋，遮瑕又會卡。呢支膠原眼霜質地輕但滋潤度夠，吸收快唔會令眼妝浮起——早晚都用得。",
        bullets=[("撫平乾紋", "眼下唔再卡遮瑕。"),
                 ("緊緻眼周", "膠原成分改善鬆弛。"),
                 ("吸收快", "唔會令眼妝浮起。"),
                 ("溫和不刺激", "眼周敏感都用得。")],
        how="早晚取米粒大小，以無名指沿眼周輕輕按壓。")),

    ("tirtir-hydra-rescue-serum", "TIRTIR 保濕急救精華", "精華", T_SKIN + ", serum",
     lambda t: "保濕急救精華" in t, dict(
        hook="皮膚鬧脾氣嗰陣用。",
        lede="換季、熬夜、用錯嘢之後皮膚會又乾又敏感。呢支集中補水同時舒緩，通常一兩日就穩定返——放喺屋企當救兵。",
        bullets=[("即時補水", "乾到緊繃都紓緩到。"),
                 ("舒緩泛紅", "敏感期都用得。"),
                 ("質地清爽", "唔會悶痘。"),
                 ("50ml 大容量", "全臉頸部都夠。")],
        how="爽膚後取適量塗於全臉，輕輕按壓至吸收。")),

    ("tirtir-uv-protection-sun-cream", "TIRTIR 紫外線防護防曬霜", "防曬", T_SKIN + ", sunscreen",
     lambda t: "防曬霜" in t, dict(
        hook="唔泛白、唔起膠。",
        lede="防曬最大問題係搽完泛白又起膠，令底妝浮起。呢支質地輕薄好推，吸收後冇白膜——所以先會日日搽得落。",
        bullets=[("唔泛白", "冇白膜，膚色自然。"),
                 ("唔起膠", "上妝唔會搓出屑。"),
                 ("清爽不焗", "夏天都搽得落。"),
                 ("50ml 大容量", "全臉頸部都夠。")],
        how="保養最後一步取適量均勻塗於全臉，出門前 15 分鐘搽。")),
]


def rows():
    wb = openpyxl.load_workbook(STOCK_XLSX)
    ws = wb.active
    h = [c.value for c in ws[1]]
    iT, iV, iQ, iP, iB = (h.index(x) for x in
                          ("Title", "Vendor", "Variant Inventory Qty",
                           "Variant Price", "Variant Barcode"))
    return [(str(r[iT]).strip(), str(r[iB]).strip(), r[iP], r[iQ] or 0)
            for r in ws.iter_rows(min_row=2, values_only=True)
            if r[iV] and "tir" in str(r[iV]).lower().replace(" ", "")]


SHADE = re.compile(r"[\[(#]\s*([0-9][0-9.]*[A-Za-z]?[^\])]*)[\])]?\s*$")


def shade_of(title):
    m = SHADE.search(title)
    if m:
        return re.sub(r"\s+", " ", m.group(1)).strip(" #])")
    m = re.search(r"\b(\d+[.\d]*[A-Za-z]?\s+[A-Za-z][\w' ]*)$", title)
    if m:
        return m.group(1).strip()
    return "單一規格"


P = {}
all_rows = rows()
used = set()
for slug, title, ptype, tags, match, copy in LINES:
    picked = [(shade_of(t), b, q) for t, b, p, q in all_rows
              if match(t) and b not in used]
    used.update(b for _, b, _ in picked)
    if not picked:
        continue
    P[slug] = dict(title=title, type=ptype, tags=tags,
                   price=max(p for t, b, p, q in all_rows if b in {x[1] for x in picked}),
                   specs=SPECS, hook=copy["hook"], lede=copy["lede"],
                   bullets=copy["bullets"], how=copy["how"],
                   shades=sorted(set(picked)))

_missed = [t for t, b, _, _ in all_rows if b not in used]
if _missed:
    print(f"未分組（{len(_missed)}）: " + "; ".join(_missed[:6]))

run(__name__, VENDOR, P, "tirtir")
