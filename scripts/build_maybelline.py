#!/usr/bin/env python3
"""
Build and publish the MAYBELLINE range.

The one brand here whose supplier titles are already Traditional Chinese,
so its HK site is wanted for imagery only. It is also the only mass-market
brand in the shop — the copy stays plainer than the K-beauty lines, since
what sells these is coverage, wear time and price rather than a texture
story.

    python3 scripts/build_maybelline.py mirror
    python3 scripts/build_maybelline.py publish [--dry-run]
"""
import re
import subprocess
import sys

import openpyxl

from brand_build import run

VENDOR = "MAYBELLINE"
STOCK_XLSX = "/Volumes/core/下載/cosmetic_products_shopify.xlsx"

T_BASE = "MAYBELLINE, 彩妝, 底妝, makeup, base"
T_EYE = "MAYBELLINE, 彩妝, 眼妝, makeup, eye"
T_LIP = "MAYBELLINE, 彩妝, 唇妝, makeup, lip"
SPECS = ["原廠正貨"]

LINES = [
    ("maybelline-fit-me-concealer", "Maybelline Fit Me 柔滑遮瑕膏", "底妝",
     T_BASE + ", concealer", lambda t: "Fit Me 柔滑遮瑕" in t, dict(
        hook="賣咗十年都仲係銷量第一。",
        lede="遮瑕膏最怕卡紋同浮粉。Fit Me 質地柔滑貼膚，眼下同痘印都遮得到，而且色階多——呢個價錢好難搵到咁齊。",
        bullets=[("柔滑貼膚", "眼下唔卡細紋。"), ("中高遮蓋", "黑眼圈同痘印都遮到。"),
                 ("色階齊全", "白皮到自然色都揀到。"), ("超值價位", "日常用唔會肉痛。")],
        how="點於瑕疵位，以指腹或遮瑕刷由中央向外輕拍。")),

    ("maybelline-instant-age-rewind-concealer", "Maybelline Cushion 遮瑕筆", "底妝",
     T_BASE + ", concealer", lambda t: "Cushion遮瑕筆" in t, dict(
        hook="海綿頭一掃，唔使工具。",
        lede="筆頭自帶海綿，轉兩下出膏、直接掃喺眼下就推得開。趕時間或者喺車上補妝最方便——唔使手指唔使刷。",
        bullets=[("自帶海綿頭", "免工具，一支完成。"), ("提亮眼下", "同時遮黑眼圈。"),
                 ("唔卡紋", "質地滋潤唔積聚。"), ("七個色階", "白皙到小麥色。")],
        how="轉出適量後直接以海綿頭點於眼下，再輕輕拍勻。")),

    ("maybelline-fit-me-matte-poreless-foundation", "Maybelline Fit Me 反孔特霧粉底液 SPF22",
     "底妝", T_BASE + ", foundation", lambda t: "反孔特霧粉底液" in t, dict(
        hook="油肌全日唔會反光。",
        lede="控油粉底通常好乾好粉。Fit Me 特霧版收毛孔同控油之餘唔會乾到見紋，而且自帶 SPF22——夏天返工啱曬。",
        bullets=[("控油特霧", "全日唔反光。"), ("柔焦毛孔", "妝面平滑。"),
                 ("SPF22", "日常通勤足夠。"), ("唔乾唔見紋", "控油唔等於乾。")],
        how="以粉底刷或海綿由面中央向外推開。")),

    ("maybelline-superstay-lumi-matte-foundation",
     "Maybelline SUPERSTAY 超持久30H空氣感粉底液", "底妝", T_BASE + ", foundation",
     lambda t: "30H空氣感粉底液" in t, dict(
        hook="30 小時，唔使補。",
        lede="質地輕如空氣但遮瑕力唔輸，而且持妝去到 30 小時——長工時、飲宴、旅行呢啲場合，唔使中途補妝。",
        bullets=[("30 小時持妝", "唔氧化唔脫妝。"), ("空氣感質地", "輕薄唔厚重。"),
                 ("中高遮蓋", "薄塗已經勻。"), ("七個色階", "冷暖膚調都有。")],
        how="以粉底刷或海綿由面中央向外推開；瑕疵位再輕拍一層。")),

    ("maybelline-superstay-creampact-foundation",
     "Maybelline 月光小忌廉 SUPERSTAY 超持久30H 氣墊粉底霜", "氣墊粉底",
     T_BASE + ", cushion", lambda t: "月光小忌廉" in t, dict(
        hook="忌廉質地，氣墊方便。",
        lede="粉底霜嘅遮瑕力配氣墊嘅方便。膏體似忌廉咁滑，輕拍就勻，30 小時持妝——想要高遮瑕又想快就揀呢個。",
        bullets=[("忌廉膏體", "滑順易推，唔會斑駁。"), ("30 小時持妝", "唔氧化唔脫妝。"),
                 ("高遮蓋", "痘印泛紅一次過。"), ("四個色階", "白皮到自然色。")],
        how="以粉撲輕拍上臉，由面中央向外推開。")),

    ("maybelline-super-stay-double-fixer-spray",
     "Maybelline SUPERSTAY 超持久24H小奶蓋定妝噴霧", "底妝", T_BASE + ", setting spray",
     lambda t: "定妝噴霧" in t, dict(
        hook="正裝連補充裝，用足好耐。",
        lede="噴完成膜，底妝唔會轉印到口罩同衣領上。24 小時定妝，而且一套有正裝加補充裝——性價比高過買兩支。",
        bullets=[("24 小時定妝", "唔溶妝唔轉印。"), ("噴完即乾", "唔黏笠。"),
                 ("連補充裝", "一套用兩次。"), ("唔改變妝感", "唔會變啞或變油。")],
        how="化妝後距離面部 20cm 均勻噴灑，待乾。")),

    ("maybelline-define-blend-brow-pencil", "Maybelline 專業柔霧造型眉筆", "眉筆",
     T_EYE + ", brow", lambda t: "眉筆" in t, dict(
        hook="一頭畫、一頭掃，眉毛即刻有形。",
        lede="筆芯扁身，可以打側填色亦可以企起描毛流；另一端螺旋刷梳順同柔化痕跡——一支完成整條眉。",
        bullets=[("扁身筆芯", "填色描線一支搞掂。"), ("附螺旋刷", "梳順柔化痕跡。"),
                 ("柔霧質感", "自然唔死板。"), ("六色可選", "由深濃啡到粉紅啡。")],
        how="順住毛流輕描，稀疏位填色，最後以螺旋刷梳勻。")),

    ("maybelline-hyper-sharp-extreme-liner", "Maybelline 超銳目極限持久眼線筆", "眼線",
     T_EYE + ", eyeliner", lambda t: "眼線筆" in t, dict(
        hook="0.01mm 筆尖，畫得出髮絲級細線。",
        lede="極細筆尖畫內眼線同眼尾都定得住，唔會斷墨要重複描。成膜後防水防油，眨極都唔會印落眼窩。",
        bullets=[("極細筆尖", "髮絲級細線都畫到。"), ("防暈防水", "唔印上眼窩。"),
                 ("一筆到底", "唔會斷墨。"), ("三色可選", "純黑、栗棕、茶棕。")],
        how="沿睫毛根部由眼頭畫向眼尾，尾段輕輕拉長。")),

    ("maybelline-sky-high-mascara", "Maybelline 飛天翹防水睫毛膏", "睫毛膏",
     T_EYE + ", mascara", lambda t: "飛天翹防水睫毛膏" in t, dict(
        hook="全球爆紅嗰支 Sky High。",
        lede="刷頭幼長，可以逐根梳到尾端，令睫毛望落長咗一截。防水配方，成日戴口罩或者夏天出汗都唔會暈。",
        bullets=[("極致纖長", "刷到尾端，長咗一截。"), ("防水防汗", "唔會暈落眼底。"),
                 ("根根分明", "唔結塊唔黐埋。"), ("兩色可選", "黑色同暮光棕。")],
        how="Z 字形由睫毛根部向外刷，重點加強眼中位置。")),

    ("maybelline-hypercurl-mascara", "Maybelline 瞬盈防水睫毛液（超鬈曲版）", "睫毛膏",
     T_EYE + ", mascara", lambda t: "瞬盈防水睫毛液" in t, dict(
        hook="彎刷頭，一刷就翹。",
        lede="刷頭做成弧形，貼住睫毛弧度一刷就撐起捲度，唔使夾都有形。防水配方，全日唔會冧返落嚟。",
        bullets=[("弧形刷頭", "貼合睫毛，一刷即翹。"), ("持久定捲", "唔使夾都撐得住。"),
                 ("防水防汗", "唔暈唔跌粉。"), ("溫水易卸", "減少甩睫毛。")],
        how="以彎刷頭貼住睫毛根部向上刷起。")),

    ("maybelline-colossal-waterproof-mascara", "Maybelline 無極限濃密睫毛液", "睫毛膏",
     T_EYE + ", mascara", lambda t: "無極限濃密睫毛液" in t, dict(
        hook="刷一次就似戴咗假睫毛。",
        lede="膏體濃密，一刷就令睫毛量感倍增。想要濃烈眼妝、或者睫毛本身稀疏，呢支比纖長型有效得多。",
        bullets=[("極致濃密", "一刷倍增量感。"), ("防水配方", "唔暈唔化。"),
                 ("大刷頭", "一次覆蓋整排睫毛。"), ("濃黑顯色", "眼神即刻有神。")],
        how="由睫毛根部向外 Z 字形刷；想更濃就等半乾後再刷一層。")),

    ("maybelline-sky-high-set", "Maybelline SKY HIGH 飛天激長睫毛 + 零NG眼線組合", "睫毛膏",
     T_EYE + ", mascara, set", lambda t: "SKY HIGH" in t.upper() and "組合" in t, dict(
        hook="眼妝兩件頭，一次過。",
        lede="Sky High 纖長睫毛膏加 HyperSharp 極細眼線筆——眼妝最關鍵嗰兩件，套裝買抵過分開。",
        bullets=[("兩件套裝", "睫毛膏＋眼線筆。"), ("皆為銷量冠軍", "各自都係熱賣款。"),
                 ("均防水", "全日唔暈。"), ("買套裝更抵", "分開買貴啲。")],
        how="先畫眼線，待乾後再刷睫毛膏。")),

    ("maybelline-superstay-vinyl-ink", "Maybelline 超持久水光唇膏液", "唇釉",
     T_LIP + ", liptint", lambda t: "超持久水光唇膏液" in t, dict(
        hook="鏡面水光，但唔會甩色。",
        lede="水光唇釉通常一飲水就冇。Vinyl Ink 成膜後色素鎖住，光澤度極高但唔黐——外出一日都唔使補。",
        bullets=[("鏡面光澤", "唇部即刻立體。"), ("超持久", "食完嘢都唔會全脫。"),
                 ("唔黐笠", "唔會黐頭髮。"), ("高顯色", "一層就夠飽和。")],
        how="以刷頭沿唇形塗一層，待成膜後再飲食。")),

    ("maybelline-lifter-plump", "Maybelline 透明質酸「嘟嘟」唇蜜", "唇彩",
     T_LIP + ", lipgloss", lambda t: "嘟嘟" in t, dict(
        hook="搽完唇會微微刺熱，然後嘟起。",
        lede="配方含辣椒素同透明質酸——刺熱感係血液循環加快，唇部自然飽滿；透明質酸同時補水，所以唔會愈搽愈乾。",
        bullets=[("即時豐盈", "唇形自然嘟起。"), ("透明質酸保濕", "唔會乾唔會起皮。"),
                 ("鏡面光澤", "唇部立體有光。"), ("四色可選", "裸調到莓紅。")],
        how="直接塗於唇部；可單搽或疊喺唇膏上。")),

    ("maybelline-eye-lip-makeup-remover", "Maybelline 眼唇二合一卸妝液", "卸妝",
     "MAYBELLINE, 護膚, skincare, cleanser", lambda t: "卸妝液" in t, dict(
        hook="防水眼妝，唔使用力擦。",
        lede="雙層配方搖勻後敷幾秒，防水睫毛膏同唇釉會自己溶開——唔使死擦眼皮，減少拉扯同甩睫毛。",
        bullets=[("溶解防水彩妝", "唔使用力擦。"), ("溫和不刺激", "眼周敏感都用得。"),
                 ("唔留油膜", "卸完唔黏笠。"), ("眼唇通用", "一支搞掂。")],
        how="搖勻後倒於化妝棉，敷於眼唇 5 秒再輕輕抹走。")),
]


def rows():
    wb = openpyxl.load_workbook(STOCK_XLSX)
    ws = wb.active
    h = [c.value for c in ws[1]]
    iT, iV, iQ, iP, iB = (h.index(x) for x in
                          ("Title", "Vendor", "Variant Inventory Qty",
                           "Variant Price", "Variant Barcode"))
    return [(str(r[iT]).strip(), str(r[iB]).strip(), r[iP], r[iQ] or 0)
            for r in ws.iter_rows(min_row=2, values_only=True)
            if r[iV] and "maybelline" in str(r[iV]).lower()]


def shade_of(title, keep):
    t = re.sub(r"^\(?o?\)?\s*(MAYBELLINE|Maybelline)\s*", "", title).strip()
    t = re.sub(re.escape(keep) + r"\s*", "", t, flags=re.I).strip(" -–")
    t = re.sub(r"\s*\d+(\.\d+)?\s*(g|ml)\s*$", "", t, flags=re.I).strip()
    return re.sub(r"\s{2,}", " ", t) or "單一規格"


P = {}
all_rows = rows()
used = set()
for slug, title, ptype, tags, match, copy in LINES:
    key = re.sub(r"^Maybelline\s*", "", title)
    picked = [(shade_of(t, key), b, q) for t, b, p, q in all_rows
              if match(t) and b not in used]
    used.update(b for _, b, _ in picked)
    if not picked:
        continue
    seen, uniq = set(), []
    for n, b, q in picked:
        n2, i = n, 2
        while n2 in seen:
            n2, i = f"{n} ({i})", i + 1
        seen.add(n2)
        uniq.append((n2, b, q))
    P[slug] = dict(title=title, type=ptype, tags=tags,
                   price=max(p for t, b, p, q in all_rows if b in {x[1] for x in uniq}),
                   specs=SPECS, hook=copy["hook"], lede=copy["lede"],
                   bullets=copy["bullets"], how=copy["how"], shades=sorted(uniq))

_missed = [t for t, b, _, _ in all_rows if b not in used]
if _missed:
    print(f"未分組（{len(_missed)}）: " + "; ".join(_missed[:5]))

if __name__ == "__main__" and len(sys.argv) > 1 and sys.argv[1] == "mirror":
    subprocess.run([sys.executable, "scripts/fetch_maybelline_hk.py"], check=True)
    subprocess.run([sys.executable, "scripts/optimise_brand_images.py",
                    "brands/maybelline"], check=True)
    sys.exit(0)

run(__name__, VENDOR, P, "maybelline")
