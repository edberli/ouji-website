#!/usr/bin/env python3
"""
Build and publish the wakemake range.

wakemake has an official Hong Kong store, so its product names are taken
from there in Traditional Chinese rather than translated from the
supplier list's mixed English. Shades, barcodes, prices and stock still
come from our own workbook.

The brand is Olive Young's own label, positioned on everyday wearability
at a low price, so the copy stays practical — what it fixes, not what it
promises.

    python3 scripts/build_wakemake.py mirror
    python3 scripts/build_wakemake.py publish [--dry-run]
"""
import re
import subprocess
import sys

import openpyxl

from brand_build import run

VENDOR = "WAKEMAKE"
STOCK_XLSX = "/Volumes/core/下載/cosmetic_products_shopify.xlsx"

T_EYE = "WAKEMAKE, K-Beauty, 彩妝, 眼妝, makeup, eye"
T_LIP = "WAKEMAKE, K-Beauty, 彩妝, 唇妝, makeup, lip"
T_CHEEK = "WAKEMAKE, K-Beauty, 彩妝, 修容, makeup, cheek"
T_BASE = "WAKEMAKE, K-Beauty, 彩妝, 底妝, makeup, base"
SPECS = ["產地：韓國 Made in Korea"]

# slug -> (HK store title, type, tags, supplier-title matcher, copy)
LINES = [
    ("wakemake-soft-blurring-eye-palette", "WAKEMAKE 柔霧經典16色眼影盤", "眼影",
     T_EYE + ", eyeshadow, palette", r"Soft Blurring Eye Palette", dict(
        hook="十六格，由返工妝一路去到派對。",
        lede="十六格入面啞光、珠光、閃片齊全，而且由淺到深排好——即係話你唔使再買第二盤來配色，一盤已經覆蓋日常同正式場合。",
        bullets=[("十六格齊全", "啞光、珠光、閃片一次過。"),
                 ("配色已諗好", "順住格數用就係完整妝容。"),
                 ("粉質綿密", "唔飛粉，唔會落喺眼底。"),
                 ("五款主題", "由裸調、玫瑰到大地色系。")],
        how="淺色打底整個眼窩，中間色暈染褶位，深色壓眼尾，閃片點眼中央。")),

    ("wakemake-healthy-glow-balm-stick", "WAKEMAKE 柔亮水潤唇膏", "唇膏",
     T_LIP + ", lipstick", r"Healthy Glow Balm Stick", dict(
        hook="潤唇膏嘅舒適度，唇膏嘅顏色。",
        lede="想有色又怕乾——呢支膏體含潤唇成分，滑過唇部即刻化開，顏色柔和有光澤，唔會勒住唇紋。趕時間唔照鏡都塗得均勻。",
        bullets=[("滋潤唔乾", "唔會起皮，唔會顯唇紋。"),
                 ("自然光澤", "水潤但唔油亮。"),
                 ("一塗即勻", "唔使唇刷，唔使照鏡。"),
                 ("五色可選", "日常裸調到玫瑰紅。")],
        how="直接以膏體由唇中央向外塗抹。")),

    ("wakemake-over-blurring-pot", "WAKEMAKE 雲朵柔霧胭脂唇霜", "胭脂",
     T_CHEEK + ", blush", r"Over Blurring Pot", dict(
        hook="唇同頰用同一盒，妝感自然統一。",
        lede="霜狀質地一觸即化，指腹輕拍就融入底妝，唔會浮喺粉底上面。同一盒點埋唇，色調自動夾——出門帶一盒就夠。",
        bullets=[("唇頰兩用", "一盒搞掂全臉氣色。"),
                 ("雲朵柔霧", "霧感但唔乾，唔會結塊。"),
                 ("免工具", "手指拍兩下就完成。"),
                 ("四色可選", "由蜜桃、玫瑰到莓調。")],
        how="指腹沾取少量，點於顴骨後由內向外輕拍；點唇時由中央推開。")),

    ("wakemake-real-defining-pencil-liner", "WAKEMAKE 柔和輕裸眼線筆", "眼線",
     T_EYE + ", eyeliner", r"Real Defining Pencil Liner", dict(
        hook="畫內眼線唔會拮親。",
        lede="筆芯柔軟順滑，畫喺睫毛根部之間唔會拉扯眼皮，但成膜後防水防油——出油同流汗都唔會印落眼窩。",
        bullets=[("柔軟筆芯", "畫內眼線唔會刺激。"),
                 ("防暈防水", "唔會印上眼窩。"),
                 ("可暈染", "落筆後三十秒內推得開。"),
                 ("五色可選", "由純黑、啡調到裸調。")],
        how="沿睫毛根部由眼頭畫向眼尾；想柔和啲就用棉花棒輕輕暈開。")),

    ("wakemake-soft-drawing-slim-brow", "WAKEMAKE 超幼柔滑持久眉筆", "眉筆",
     T_EYE + ", brow", r"3 in 1", dict(
        hook="一支三用：畫、填、梳。",
        lede="幼筆芯描得出單根毛流，另一端嘅螺旋刷梳順同柔化痕跡，中間仲有暈染頭填稀疏位——一支就完成整條眉。",
        bullets=[("超幼筆芯", "描得出單根毛流。"),
                 ("三合一", "畫、填、梳一支做齊。"),
                 ("防水持久", "全日唔甩色。"),
                 ("五色貼近髮色", "由深啡到灰棕。")],
        how="順住毛流一條條輕描，用暈染頭填稀疏位，最後以螺旋刷梳勻。")),

    ("wakemake-defining-cover-concealer", "WAKEMAKE 輕透無瑕遮瑕膏 SPF30 PA++", "底妝",
     T_BASE + ", concealer", r"輕透無瑕遮瑕膏", dict(
        hook="遮到黑眼圈，但唔會卡紋。",
        lede="眼下遮瑕最怕落粉之後全部卡喺細紋度。呢支質地輕薄貼膚，遮蓋力夠但唔會積聚，而且自帶 SPF30——遮瑕同防曬一步做齊。",
        bullets=[("唔卡細紋", "眼下用一整日都平滑。"),
                 ("自帶 SPF30", "慳一個步驟。"),
                 ("高遮蓋", "黑眼圈同痘印都遮得到。"),
                 ("四個色階", "由白皮到自然色。")],
        how="點於瑕疵位，以指腹或遮瑕刷由中央向外輕拍。")),

    ("wakemake-defining-cover-conceal-fit-palette", "WAKEMAKE 輕透無瑕遮瑕修容盤", "底妝",
     T_BASE + ", concealer, palette", r"輕透無瑕遮瑕修容盤", dict(
        hook="遮瑕、校色、修容，一盤搞掂。",
        lede="唔同瑕疵要唔同顏色去遮：泛紅要綠調、黑眼圈要蜜桃調。呢盤把校色同遮瑕格擺埋一齊，可以按位置調配，唔使買幾支。",
        bullets=[("多格配色", "校色同遮瑕按需要調配。"),
                 ("貼膚唔厚重", "唔會有面具感。"),
                 ("旅行啱用", "一盤頂幾支。"),
                 ("兩款可選", "跟膚色深淺揀。")],
        how="以遮瑕刷沾取對應色，點於瑕疵位輕拍推開。")),

    ("wakemake-seamless-wear-foundation", "WAKEMAKE 柔焦裸肌無瑕粉底液", "底妝",
     T_BASE + ", foundation", r"Seamless Wear Foundation", dict(
        hook="遮得到，但望落仲係皮膚。",
        lede="遮瑕力同自然感通常二選一。呢支質地輕薄但色素密，薄塗已經勻，柔焦收尾令毛孔同細紋淡化，唔使為咗遮瑕搽到似面具。",
        bullets=[("薄塗高遮", "唔使疊厚。"),
                 ("柔焦毛孔", "妝面平滑均勻。"),
                 ("持妝唔氧化", "全日唔變深變黃。"),
                 ("三個色階", "白皮到自然色。")],
        how="以粉底刷或海綿由面中央向外推開；瑕疵位再輕拍一層。")),

    ("wakemake-water-glow-coating-cushion", "WAKEMAKE 水光亮感貼肌氣墊 SPF50+ PA++++",
     "氣墊粉底", T_BASE + ", cushion", r"水光亮感貼肌氣墊", dict(
        hook="水光底妝，加最高防曬。",
        lede="薄薄一層就夠勻，妝感係健康皮膚嘅光澤而唔係粉底嘅厚度，而且做到 SPF50+ PA++++——夏天出街唔使另外再搽防曬。",
        bullets=[("SPF50+ PA++++", "最高等級防曬。"),
                 ("水光貼膚", "亮而唔油，唔會似出汗。"),
                 ("薄塗夠遮", "唔使疊厚就均勻。"),
                 ("三個色階", "白皮到自然色。")],
        how="以粉撲輕拍上臉，由面中央向外推開。")),

    ("wakemake-real-defining-brush-liner", "WAKEMAKE 柔和輕裸眼線液", "眼線",
     T_EYE + ", eyeliner", r"Real Defining Brush Liner", dict(
        hook="極細筆尖，眼尾拉得準。",
        lede="毛刷筆尖夠幼夠挺，畫細線同拉眼尾都定得住，唔會斷墨要重複描。成膜後防水防油，眨極都唔會印上眼窩。",
        bullets=[("極細筆尖", "眼尾同內眼線都畫得準。"),
                 ("一筆到底", "唔會斷墨。"),
                 ("防暈防水", "唔印上眼窩。"),
                 ("三色可選", "純黑、啡調同裸調。")],
        how="沿睫毛根部由眼頭畫向眼尾，尾段輕輕拉長。")),

    ("wakemake-real-defining-lash-mascara", "WAKEMAKE 輕透無重睫毛膏", "睫毛膏",
     T_EYE + ", mascara", r"Real Defining Lash Mascara", dict(
        hook="刷完唔會覺得眼皮重。",
        lede="睫毛膏最無奈係刷完睫毛頂唔住，中午就冧返落嚟。呢支膏體輕，根根分明唔結塊，捲度由根部撐住成日。",
        bullets=[("輕盈唔重墜", "捲度企得住。"),
                 ("根根分明", "唔結塊唔黐埋。"),
                 ("唔暈唔跌粉", "眼底唔會出現黑影。"),
                 ("溫水易卸", "減少甩睫毛。")],
        how="Z 字形由睫毛根部向外刷，重點加強眼中位置。")),

    ("wakemake-stay-fixer-multi-color-powder", "WAKEMAKE 3色亮膚定妝蜜粉", "底妝",
     T_BASE + ", powder", r"Stay Fixer Multi Color Powder", dict(
        hook="定妝同校色，一次過。",
        lede="三種色調分別壓泛紅、暗黃同灰暗，可以單獨用喺對應位置，亦可以掃勻做整體提亮。粉體極細，定妝之後摸落仲係滑嘅。",
        bullets=[("三色校色", "分區處理泛紅同暗黃。"),
                 ("極細粉體", "唔見紋，唔會浮粉。"),
                 ("控油唔乾", "定妝之餘保留潤度。"),
                 ("兩款可選", "跟膚色揀。")],
        how="以蜜粉刷沾取，掃勻全臉；亦可單取一色用喺需要校色嘅位置。")),

    ("wakemake-vitamin-tone-up-lotion", "WAKEMAKE 維他命水嫩光感底霜 SPF", "底妝",
     T_BASE + ", primer", r"維他命水嫩光感底霜", dict(
        hook="提亮唔死白。",
        lede="一般提亮底霜會令膚色變得灰白。呢款靠維他命成分提亮，膚色自然透亮之餘同時保濕——底妝上去會貼服好多。",
        bullets=[("自然提亮", "唔會死白。"),
                 ("同時保濕", "乾燥位唔會起皮。"),
                 ("平滑底妝", "粉底更均勻更貼。"),
                 ("含防曬", "日常通勤足夠。")],
        how="潔面保養後取適量薄塗全臉，待吸收後再上底妝。")),
]


def rows():
    wb = openpyxl.load_workbook(STOCK_XLSX)
    ws = wb.active
    h = [c.value for c in ws[1]]
    iT, iV, iQ, iP, iB = (h.index(x) for x in
                          ("Title", "Vendor", "Variant Inventory Qty",
                           "Variant Price", "Variant Barcode"))
    return [(str(r[iT]).strip(), str(r[iB]).strip(), r[iP], r[iQ] or 0)
            for r in ws.iter_rows(min_row=2, values_only=True)
            if r[iV] and "wakemake" in str(r[iV]).lower()]


def shade_of(title):
    m = re.search(r"\[(.+?)\]\s*$", title)
    if m:
        return m.group(1).strip()
    m = re.search(r"(\d+[A-Za-z]?\s+[^\d]+)$", title)
    return m.group(1).strip() if m else title


P = {}
all_rows = rows()
used = set()
for slug, title, ptype, tags, match, copy in LINES:
    picked = [(shade_of(t), b, q, p) for t, b, p, q in all_rows
              if re.search(match, t, re.I) and b not in used]
    used.update(b for _, b, _, _ in picked)
    if not picked:
        continue
    P[slug] = dict(title=title, type=ptype, tags=tags,
                   price=max(p for _, _, _, p in picked) or 0, specs=SPECS,
                   hook=copy["hook"], lede=copy["lede"],
                   bullets=copy["bullets"], how=copy["how"],
                   shades=sorted((n, b, q) for n, b, q, _ in picked))

_missed = [t for t, b, _, _ in all_rows if b not in used]
if _missed:
    print(f"未分組（{len(_missed)}）: " + "; ".join(_missed[:6]))

if __name__ == "__main__" and len(sys.argv) > 1 and sys.argv[1] == "mirror":
    subprocess.run([sys.executable, "scripts/fetch_wakemake_hk.py"], check=True)
    subprocess.run([sys.executable, "scripts/optimise_brand_images.py",
                    "brands/wakemake"], check=True)
    sys.exit(0)

run(__name__, VENDOR, P, "wakemake")
