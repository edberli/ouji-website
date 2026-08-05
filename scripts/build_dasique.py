#!/usr/bin/env python3
"""
Build and publish the dasique range.

dasique.com is a Shopify store, so imagery comes straight off its
products.json — but it carries no barcodes, and its numbering is the only
thing our supplier titles share with it. Every one of our SKUs is
"<line> #<n> <中文色名>", so a line is matched by name and a shade by its
number inside that line.

Two shapes on their side, handled the same way:

  * a multi-variant product per line (Pure Water Lip Gloss 01–14) —
    images are the line's, shared by every shade
  * one product per shade (Eyeshadow Palette - 24 Muted Nuts) — images
    are that shade's alone, which is better and is preferred when both
    exist

A line whose shades match nothing keeps its copy and ships as a draft
rather than going live with no picture.

    python3 scripts/build_dasique.py --dry-run
    python3 scripts/build_dasique.py
"""
import argparse
import json
import os
import re
import sys
import urllib.request

import openpyxl

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from publish import publish  # noqa: E402

STORE = "https://dasique.com"
UA = ("Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 "
      "(KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36")
STOCK_XLSX = "/Volumes/core/下載/cosmetic_products_shopify.xlsx"
VENDOR = "dasique"

T_EYE = "dasique, 韓國彩妝, 彩妝, 眼妝, K-Beauty, makeup, eye"
T_LIP = "dasique, 韓國彩妝, 彩妝, 唇妝, K-Beauty, makeup, lip"
T_CHEEK = "dasique, 韓國彩妝, 彩妝, 修容, K-Beauty, makeup, cheek"
T_BASE = "dasique, 韓國彩妝, 彩妝, 底妝, K-Beauty, makeup, base"

# our title keyword -> (handle, our line name, type, tags, store line
# titles to pull imagery from, copy)
LINES = [
    # Longer names first: "眼影盤" is a substring of both of these, so
    # they have to claim their rows before the plain palette line does.
    ("二十色眼影", "dasique-20-color-eyeshadow-palette", "dasique 二十色眼影盤", "眼影",
     T_EYE + ", eyeshadow, palette", [], dict(
        hook="二十格，一盤打天下。",
        lede="想一個盤玩到日常、約會同派對，就要格數夠多。二十色由裸色到深調、由啞光到亮片全部齊，唔使再為咗一隻顏色買多個盤。",
        bullets=[("二十格色域", "一盤搞掂所有場合。"),
                 ("質地齊全", "啞光、珠光、亮片一次過。"),
                 ("疊色友好", "深淺之間過渡順，唔會邊界明顯。"),
                 ("旅行啱用", "帶一個盤就夠。")],
        how="先用裸色打底，再按當日妝感揀啞光或亮片疊上去。")),

    ("星夜眼影盤", "dasique-starlit-jewel-set", "dasique 星夜眼影四件組", "眼影",
     # Starlit Jewel on their store is the liquid glitter, not this
     # four-piece set, and the shade names do not line up — better a
     # draft than the wrong product's photos.
     T_EYE + ", eyeshadow, glitter, set", [], dict(
        hook="四件組，閃片係主角。",
        lede="液態閃片上眼會貼實而唔會飛落面，所以派對妝可以放心用。四件一組配好色，唔使再逐支揀。",
        bullets=[("液態閃片", "貼實唔飛粉。"),
                 ("四件成組", "配色已經夾好。"),
                 ("易卸", "唔會拮眼難卸。"),
                 ("節日啱用", "派對妝一步到位。")],
        how="以指腹或扁刷點喺眼中央，唔使暈開。")),

    ("眼影盤", "dasique-eyeshadow-palette", "dasique 九色眼影盤", "眼影",
     T_EYE + ", eyeshadow, palette", ["Eyeshadow Palette"], dict(
        hook="一盤九色，配色已經幫你配好。",
        lede="dasique 最出名就係呢啲盤——啞光、微閃、亮片排得好，順住格數用落去就係一個完整眼妝，唔使自己諗邊隻夾邊隻。粉質細滑唔飛粉，新手都唔會落手落腳整污糟。",
        bullets=[("配色唔使諗", "由淺到深排好，順住用就得。"),
                 ("三種質地", "啞光打底、微閃疊層、亮片點睛。"),
                 ("粉質細滑", "顯色但唔飛粉，唔會愈掃愈污。"),
                 ("色號夠齊", "由日常大地到粉紫莓調都有。")],
        how="淺色打底，中間色暈染褶位，深色壓眼尾，亮片點喺眼中央。")),

    ("腮紅盤", "dasique-blending-mood-cheek", "dasique 混色腮紅盤", "胭脂",
     T_CHEEK + ", blush, palette", ["Blending Mood Cheek"], dict(
        hook="四格腮紅，自己調到啱色。",
        lede="單色腮紅最麻煩就係唔啱膚色又救唔返。呢個盤一格一格可以自己撈，深咗就掃返淺色沖淡，冷暖都夾到——所以四季都用得着。",
        bullets=[("四格自由調", "撈到啱自己膚色為止。"),
                 ("粉質軟糯", "貼膚唔浮粉，唔會一撻撻。"),
                 ("顯色好控", "逐層疊，唔會一下太紅。"),
                 ("同眼影同名", "配對應色號眼影盤，全臉一次過。")],
        how="以腮紅掃沾單格或撈兩格，喺手背拍走多餘粉量，再由笑肌向太陽穴輕掃。")),

    ("高光盤", "dasique-shine-glowy-highlighter-palette", "dasique 貝殼高光盤", "高光",
     T_CHEEK + ", highlighter", ["Shine Glowy Highlighter Palette"], dict(
        hook="細閃唔飛粉，唔會變油光。",
        lede="高光最怕閃片大粒，燈光一照就好似出油。呢個盤粉體極細，上面係一層由皮膚透出嚟嘅光而唔係一層粉——影相唔會反白。",
        bullets=[("極細閃粉", "自然光澤，唔會粒粒可見。"),
                 ("多格可疊", "淡光同強光自己揀。"),
                 ("貼膚唔飛粉", "唔會落喺毛孔位。"),
                 ("眼下都用得", "點喺臥蠶提亮眼神。")],
        how="以細掃沾取，點喺顴骨上緣、鼻樑同唇珠。")),

    ("遮瑕盤", "dasique-pro-concealer-palette", "dasique 專業遮瑕盤", "遮瑕",
     T_BASE + ", concealer, palette", ["Pro Concealer Palette"], dict(
        hook="校色同遮瑕分開嚟做。",
        lede="黑眼圈同泛紅唔應該用同一隻色去冚。一盤有校色同遮瑕兩類，按位置揀色再疊薄薄一層，比起厚塗一隻遮瑕自然好多。",
        bullets=[("校色＋遮瑕", "唔同瑕疵用唔同格。"),
                 ("膏體軟身", "體溫一推就化開，唔會卡紋。"),
                 ("薄塗夠力", "唔使厚搽都遮到。"),
                 ("兩款可選", "跟膚色深淺揀。")],
        how="以遮瑕刷沾取，點喺瑕疵位由中心向外輕拍推開。")),

    ("果汁鏡面唇釉", "dasique-juicy-dewy-lip-tint", "dasique 果汁鏡面唇釉", "唇釉",
     T_LIP + ", liptint", ["Juicy Dewy Lip Tint"], dict(
        hook="鏡面水光，唔黐笠。",
        lede="dasique 嘅招牌唇釉。質地係水感而唔係糖漿，上唇薄薄一層就有鏡面反光，唔會拉絲又唔會黐住頭髮——夏天戴住都唔覺得侷。",
        bullets=[("鏡面水光", "反光度高但唔油膩。"),
                 ("唔黐唔拉絲", "唔會黐頭髮。"),
                 ("薄透顯色", "疊多層可以加深。"),
                 ("色號好日常", "果調為主，黃皮白皮都夾。")],
        how="由唇中央向外推開，想深色就再疊一層。")),

    ("水露唇蜜", "dasique-pure-water-lip-gloss", "dasique 純水唇蜜", "唇蜜",
     T_LIP + ", lipgloss", ["Pure Water Lip Gloss"], dict(
        hook="唇蜜嘅光澤，冇唇蜜嘅黐。",
        lede="一般唇蜜靚得一陣，之後就黐到唔想講嘢。呢支質地似水多過似油，光澤度夠但唔會黐——單搽或者疊喺唇釉上面加光都得。",
        bullets=[("水感質地", "唔黐唔重。"),
                 ("透明感光澤", "唔會遮走底下嘅唇色。"),
                 ("滋潤唔乾", "冇一般唇蜜嘅緊繃感。"),
                 ("可疊搽", "疊喺唇釉上面即刻加光。")],
        how="單搽於唇部，或疊於唇釉之上。")),

    ("晨露唇釉", "dasique-juicy-dewy-glow-tint", "dasique 晨露光澤唇釉", "唇釉",
     T_LIP + ", liptint", ["Juicy Dewy Glow Tint"], dict(
        hook="似朝早嗰陣露水嘅光。",
        lede="比鏡面唇釉再柔一級——唔係一層硬光，而係由唇入面透出嚟嘅水潤感。顯色度中等，所以唔會有明顯唇線，適合日常同返工。",
        bullets=[("柔和水光", "唔係硬反光，自然好多。"),
                 ("唔顯唇紋", "薄塗都唔會積喺紋度。"),
                 ("中等顯色", "唔會太搶，日常戴得住。"),
                 ("滋潤感強", "唔會愈搽愈乾。")],
        how="由唇中央向外推開，唇紋深可先塗潤唇膏打底。")),

    ("果醬唇蜜", "dasique-fruity-lip-jam", "dasique 果醬唇蜜", "唇蜜",
     T_LIP + ", lipgloss", ["Fruity Lip Jam"], dict(
        hook="果醬質地，飽和又水潤。",
        lede="比一般唇蜜濃稠，所以顯色度高好多，但因為係果醬質地所以唔會乾。想一支搞掂顯色同光澤就係呢支。",
        bullets=[("高顯色", "一層就夠飽和。"),
                 ("果醬質感", "濃稠但唔黐笠。"),
                 ("水潤持久", "唔會中途變乾起皮。"),
                 ("果調色系", "由蜜桃到無花果都有。")],
        how="以刷頭沿唇形塗勻，想清爽啲可以用手指拍散。")),

    ("愛心多用膏", "dasique-souffle-color-pot", "dasique 舒芙蕾多用彩膏", "多用彩妝",
     T_CHEEK + ", multi, blush", ["Souffle Color Pot"], dict(
        hook="唇、頰、眼一盒搞掂。",
        lede="膏狀質地一撻上面就融開，唔似粉狀咁會突顯乾紋。同一隻色用喺唇同頰上面，全臉自然統一——化妝袋唔使再帶三件嘢。",
        bullets=[("一膏三用", "唇、腮紅、眼影都得。"),
                 ("舒芙蕾質地", "體溫一推就化。"),
                 ("唔顯乾紋", "膏狀貼膚，唔會卡粉。"),
                 ("細盒易帶", "放銀包都唔佔位。")],
        how="用手指沾少量，點喺笑肌或唇上再向外拍散。")),

    ("布丁多用膏", "dasique-chewing-glow-pot", "dasique 布丁光澤多用膏", "多用彩妝",
     T_CHEEK + ", multi, blush", ["Chewing Glow Pot"], dict(
        hook="布丁質地，帶少少光。",
        lede="同舒芙蕾款一樣係膏狀，但多咗一層水光。想面部有果凍感又唔想搽多一層高光，用呢個一步搞掂。",
        bullets=[("布丁質地", "彈手，推開即化。"),
                 ("自帶水光", "唔使再搽高光。"),
                 ("唇頰同用", "全臉色調統一。"),
                 ("色域夠闊", "由蜜桃到可可莓調。")],
        how="手指沾少量點喺笑肌，向外拍散；唇上直接點塗。")),

    ("按壓式糖果滋潤唇膏", "dasique-candy-rolling-pot", "dasique 按壓式糖果潤唇膏", "潤唇膏",
     T_LIP + ", lipbalm", ["Candy Rolling Pot"], dict(
        hook="按一下就出，唔使挖。",
        lede="罐裝潤唇膏最麻煩就係要用手指挖。呢支按一下就出啱啱好嘅份量，出街用唔怕污糟——滋潤度夠之餘帶少少色。",
        bullets=[("按壓出膏", "唔使用手指挖。"),
                 ("滋潤帶色", "護唇同顯色一次過。"),
                 ("唔黐唔油", "上唇薄透。"),
                 ("細支易帶", "放袋都唔佔位。")],
        how="按壓底部出膏，直接塗於唇部。")),

    ("水光氣墊", "dasique-glow-cushion", "dasique 愛心水光氣墊", "氣墊粉底",
     T_BASE + ", cushion", [], dict(
        hook="水光底妝，唔會假白。",
        lede="妝感係濕潤有光而唔係一層粉，所以近距離睇都似皮膚本身好。愛心造型粉盒補妝拎出嚟都好睇。",
        bullets=[("水光妝感", "唔會乾唔會粉感。"),
                 ("薄透遮瑕", "遮到但唔厚重。"),
                 ("貼膚持久", "唔易斑駁。"),
                 ("愛心粉盒", "補妝拎出嚟都靚。")],
        how="以粉撲沾取，由面中央向外輕拍。")),

    ("睫毛膏", "dasique-volume-curl-mascara", "dasique 纖長捲翹睫毛膏", "睫毛膏",
     T_EYE + ", mascara", [], dict(
        hook="夾完唔會跌返落嚟。",
        lede="刷頭幼細,所以夾得起下睫毛同眼頭嗰啲短毛。膏體唔會結塊，一刷一刷分明——最緊要係定型力夠，落妝前都仲捲住。",
        bullets=[("持久捲翹", "一日都唔會塌。"),
                 ("根根分明", "唔會黐成一撻。"),
                 ("幼細刷頭", "眼頭眼尾都刷到。"),
                 ("防暈染", "唔會落熊貓眼。")],
        how="由睫毛根部向上以 Z 字掃出，分兩層加密。")),

    ("冰淇淋系列", "dasique-ice-cream-lip-tint", "dasique 冰淇淋系列唇釉", "唇釉",
     T_LIP + ", liptint", ["Ice Cream Collection"], dict(
        hook="雪糕色系，甜而唔膩。",
        lede="果汁鏡面唇釉嘅雪糕限定色，色調比正代更柔和帶奶感，上唇冇明顯唇線——想要嗰種「唇本身就係呢個色」嘅感覺就用呢支。",
        bullets=[("奶調色系", "柔和唔搶。"),
                 ("鏡面水光", "同正代一樣唔黐。"),
                 ("薄透好推", "唔會有明顯邊界。"),
                 ("限定配色", "同正代唔撞色。")],
        how="由唇中央向外推開，想加深就疊多一層。")),
]


def store_products():
    req = urllib.request.Request(f"{STORE}/products.json?limit=250",
                                 headers={"User-Agent": UA})
    return [p for p in json.load(urllib.request.urlopen(req, timeout=60))["products"]
            if not p["title"].startswith("🎁")]


def our_rows():
    wb = openpyxl.load_workbook(STOCK_XLSX)
    ws = wb.active
    h = [c.value for c in ws[1]]
    iT, iV, iQ, iP, iB = (h.index(x) for x in
                          ("Title", "Vendor", "Variant Inventory Qty",
                           "Variant Price", "Variant Barcode"))
    return [(str(r[iT]).strip(), str(r[iB]).strip(), r[iP], r[iQ] or 0)
            for r in ws.iter_rows(min_row=2, values_only=True)
            if r[iV] and "dasique" in str(r[iV]).lower()]


def shade_of(title, keyword):
    """Everything after the line name: '#13 香芋奶茶', '01 黑色', '#21C 陶瓷白'."""
    tail = title.split(keyword, 1)[-1]
    tail = re.sub(r"^[\s?？]*", "", tail).strip(" ()（）[]")
    return re.sub(r"\s{2,}", " ", tail).strip() or "單一規格"


def shade_no(shade):
    """The leading number, which is the only thing both sides share."""
    m = re.match(r"#?\s*(\d{1,2})", shade)
    return int(m.group(1)) if m else None


def images_for(store, line_titles, no):
    """Prefer the one-product-per-shade listing, since its photos are of
    that shade; fall back to the line's shared gallery."""
    if not line_titles:
        return []
    hits = [p for p in store if any(t.lower() in p["title"].lower()
                                    for t in line_titles)]
    if no is not None:
        exact = [p for p in hits
                 if re.search(rf"(?<!\d){no:02d}(?!\d)", p["title"])
                 or re.search(rf"(?<!\d){no}(?!\d)", p["title"])]
        per_shade = [p for p in exact if len(p["variants"]) == 1]
        if per_shade:
            return [i["src"] for i in per_shade[0]["images"]]
    multi = sorted([p for p in hits if len(p["variants"]) > 1],
                   key=lambda p: -len(p["images"]))
    return [i["src"] for i in multi[0]["images"]] if multi else []


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--dry-run", action="store_true")
    args = ap.parse_args()

    store = store_products()
    rows = our_rows()
    used = set()

    for keyword, handle, title, ptype, tags, line_titles, copy in LINES:
        ours = [(shade_of(t, keyword), b, q, p) for t, b, p, q in rows
                if keyword in t and b not in used]
        if not ours:
            print(f"  ?? {title}: 冇對應 SKU")
            continue
        used.update(b for _, b, _, _ in ours)

        # de-duplicate shade labels; productSet rejects a repeated option
        seen, shades = set(), []
        for name, b, q, p in ours:
            label = name
            n = 2
            while label in seen:
                label, n = f"{name} ({n})", n + 1
            seen.add(label)
            shades.append({"name": label, "barcode": b, "qty": q})

        per_shade = {s["name"]: images_for(store, line_titles,
                                           shade_no(s["name"])) for s in shades}
        imgs = list(dict.fromkeys([u for s in shades for u in per_shade[s["name"]]]))

        body = (f'<p><strong>{copy["hook"]}</strong></p><p>{copy["lede"]}</p>'
                + "<ul>" + "".join(f"<li><strong>{t}</strong>——{d}</li>"
                                   for t, d in copy["bullets"]) + "</ul>"
                + f'<p><strong>用法</strong><br>{copy["how"]}</p>')
        if len(imgs) > 1:
            body += ('<div class="product-detail-images">'
                     + "".join(f'<img src="{u}" alt="{title} 產品介紹" loading="lazy">'
                               for u in imgs[1:])
                     + "</div>")

        item = {
            "handle": handle, "title": title, "descriptionHtml": body,
            "vendor": VENDOR, "productType": ptype,
            "tags": [x.strip() for x in tags.split(",")],
            "status": "ACTIVE" if imgs else "DRAFT",
            "option_name": "色號",
            "price": max(p for _, _, _, p in ours),
            "images": imgs[:40],
            "shades": shades,
        }
        flag = "" if imgs else "   ← 冇圖，出 draft"
        print(f'{len(shades):>2} 色  {len(imgs):>2} 圖  {title}{flag}')
        if not args.dry_run:
            r = publish(item)
            print(f"        -> {r['handle']}  {r['variants']} variants, "
                  f"{r['media']} media, {r['channels']} channels")

    left = [t for t, b, _, _ in rows if b not in used]
    if left:
        print(f"\n未歸類（{len(left)}）：" + "; ".join(left))


if __name__ == "__main__":
    main()
